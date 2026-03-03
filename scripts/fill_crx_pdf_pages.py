#!/usr/bin/env python3
"""
Populate missing CRX PDF page numbers in an Excel workbook.

Workflow:
1. Read CRX rows where page column is blank and comments contain "Should be:" text.
2. Detect chapter start pages in the PDF (exact match first, fuzzy fallback).
3. Match each row's "Should be" text inside that chapter's page window.
4. Write page numbers to the workbook (or print-only in --dry-run mode).

Typical usage:
  uv run --with openpyxl --with pypdf --with rapidfuzz \
    python scripts/fill_crx_pdf_pages.py \
      --xlsx "E:/Audiobooks/Raws/Book/CRX/My_CRX.xlsx" \
      --pdf "E:/Audiobooks/Raws/Book/My-Print.pdf" \
      --book-index "E:/Audiobooks/Raws/Book/book-index.json"
"""

from __future__ import annotations

import argparse
import datetime as dt
import json
import re
import shutil
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, Iterable, List, Tuple

from openpyxl import load_workbook
from openpyxl.utils.cell import column_index_from_string
from pypdf import PdfReader
from rapidfuzz import fuzz


CHAPTER_TITLE_RE = re.compile(r"(?i)^\s*chapter\s+(\d+)\s*:\s*(.+?)\s*$")
CHAPTER_NUMBER_RE = re.compile(r"(?i)\bchapter\s+(\d+)\b")
SHOULD_BE_RE = re.compile(r"(?is)should\s+be:\s*(.*?)(?:\n\s*read\s+as:|\Z)")


@dataclass(frozen=True)
class ChapterTitle:
    number: int
    title: str
    subtitle: str


@dataclass(frozen=True)
class RowCandidate:
    row: int
    error_id: str
    chapter_label: str
    chapter_number: int
    should_text: str
    should_norm: str


@dataclass(frozen=True)
class RowMatch:
    candidate: RowCandidate
    page: int
    method: str
    score: float
    range_start: int
    range_end: int
    top_candidates: Tuple[Tuple[int, float], ...]

    @property
    def is_exact(self) -> bool:
        return self.method.startswith("exact")


def normalize_text(text: str) -> str:
    text = text or ""
    text = (
        text.replace("\u2019", " ")
        .replace("\u2018", " ")
        .replace("'", " ")
        .replace("\u201c", '"')
        .replace("\u201d", '"')
        .replace("&", " and ")
        .replace("[", " ")
        .replace("]", " ")
    )
    text = text.lower()
    text = re.sub(r"[^a-z0-9\s]+", " ", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def extract_should_be(comments: str) -> str:
    match = SHOULD_BE_RE.search(comments or "")
    if not match:
        return ""
    return match.group(1).strip()


def parse_chapter_title(title: str) -> ChapterTitle | None:
    match = CHAPTER_TITLE_RE.match(title or "")
    if not match:
        return None
    number = int(match.group(1))
    subtitle = match.group(2).strip()
    return ChapterTitle(number=number, title=title.strip(), subtitle=subtitle)


def parse_chapter_number(label: str) -> int | None:
    match = CHAPTER_NUMBER_RE.search(label or "")
    if not match:
        return None
    return int(match.group(1))


def read_pdf_pages(pdf_path: Path) -> List[str]:
    reader = PdfReader(str(pdf_path))
    pages: List[str] = []
    for page in reader.pages:
        pages.append(normalize_text(page.extract_text() or ""))
    return pages


def load_book_index_chapters(book_index_path: Path) -> Dict[int, ChapterTitle]:
    with book_index_path.open("r", encoding="utf-8") as handle:
        payload = json.load(handle)

    chapters: Dict[int, ChapterTitle] = {}
    for section in payload.get("sections", []):
        parsed = parse_chapter_title(str(section.get("title", "")).strip())
        if parsed is None:
            continue
        chapters.setdefault(parsed.number, parsed)
    return chapters


def collect_sheet_chapters(labels: Iterable[str]) -> Dict[int, ChapterTitle]:
    chapters: Dict[int, ChapterTitle] = {}
    for label in labels:
        parsed = parse_chapter_title(label)
        if parsed is None:
            continue
        chapters.setdefault(parsed.number, parsed)
    return chapters


def query_variants(chapter: ChapterTitle) -> List[str]:
    subtitle_norm = normalize_text(chapter.subtitle)
    if not subtitle_norm:
        return [normalize_text(f"{chapter.number}")]

    tokens = subtitle_norm.split()
    variants = [
        normalize_text(f"{chapter.number} {subtitle_norm}"),
        normalize_text(f"{chapter.number} {' '.join(tokens[:2])}"),
        normalize_text(f"{chapter.number} {tokens[0]}"),
    ]
    deduped: List[str] = []
    for item in variants:
        if item and item not in deduped:
            deduped.append(item)
    return deduped


def find_exact_heading_page(
    query: str, page_texts: List[str], previous_page: int | None
) -> Tuple[int | None, str]:
    if not query:
        return None, "no-query"

    hits = [idx + 1 for idx, text in enumerate(page_texts) if query in text]
    if len(hits) == 1:
        return hits[0], "exact"

    if previous_page is not None:
        ordered = [page for page in hits if page > previous_page]
        if ordered:
            return ordered[0], "exact-ordered"

    return None, "not-found"


def detect_chapter_starts(
    chapters: Dict[int, ChapterTitle],
    page_texts: List[str],
    chapter_min_confidence: float,
) -> Dict[int, Tuple[int, str, float]]:
    starts: Dict[int, Tuple[int, str, float]] = {}
    previous_page: int | None = None

    for number in sorted(chapters):
        chapter = chapters[number]
        page: int | None = None
        method = "none"
        score = 0.0

        for variant in query_variants(chapter):
            page, reason = find_exact_heading_page(variant, page_texts, previous_page)
            if page is not None:
                method = reason
                score = 100.0
                break

        if page is None:
            # Fuzzy fallback when exact heading line did not normalize cleanly.
            best_score = -1.0
            best_page = 1
            best_variant = ""
            for variant in query_variants(chapter):
                for idx, text in enumerate(page_texts, start=1):
                    current = float(fuzz.partial_ratio(variant, text))
                    if current > best_score:
                        best_score = current
                        best_page = idx
                        best_variant = variant
            if best_score >= chapter_min_confidence:
                page = best_page
                method = f"fuzzy:{best_variant}"
                score = best_score

        if page is not None:
            starts[number] = (page, method, score)
            previous_page = page

    return starts


def build_chapter_ranges(
    chapter_starts: Dict[int, Tuple[int, str, float]],
    page_count: int,
) -> Dict[int, Tuple[int, int]]:
    ranges: Dict[int, Tuple[int, int]] = {}
    ordered = sorted(chapter_starts.items(), key=lambda item: item[0])
    for index, (chapter_number, (start_page, _, _)) in enumerate(ordered):
        if index + 1 < len(ordered):
            next_start = ordered[index + 1][1][0]
            end_page = max(start_page, next_start - 1)
        else:
            end_page = page_count
        ranges[chapter_number] = (start_page, end_page)
    return ranges


def find_row_candidates(
    sheet,
    *,
    start_row: int,
    error_col: int,
    chapter_col: int,
    page_col: int,
    comments_col: int,
    force: bool,
) -> List[RowCandidate]:
    rows: List[RowCandidate] = []
    for row in range(start_row, sheet.max_row + 1):
        error_id = sheet.cell(row=row, column=error_col).value
        chapter_label = sheet.cell(row=row, column=chapter_col).value
        page_value = sheet.cell(row=row, column=page_col).value
        comments = sheet.cell(row=row, column=comments_col).value

        if not error_id or not isinstance(comments, str):
            continue

        if "should be:" not in comments.lower():
            continue

        if not force and page_value is not None and str(page_value).strip():
            continue

        chapter_number = parse_chapter_number(str(chapter_label or ""))
        if chapter_number is None:
            continue

        should_text = extract_should_be(comments)
        should_norm = normalize_text(should_text)
        if not should_norm:
            continue

        rows.append(
            RowCandidate(
                row=row,
                error_id=str(error_id).strip(),
                chapter_label=str(chapter_label or "").strip(),
                chapter_number=chapter_number,
                should_text=should_text,
                should_norm=should_norm,
            )
        )

    return rows


def longest_anchor_page_hit(
    should_norm: str,
    page_texts: List[str],
    page_range: Tuple[int, int],
) -> Tuple[int | None, str]:
    tokens = should_norm.split()
    if len(tokens) < 4:
        return None, "none"

    start_page, end_page = page_range

    # Pass 1: sentence prefix anchors (best for split-page lines).
    # Pass 2: middle/end anchors.
    for pass_index in (1, 2):
        for window in (14, 12, 10, 8, 6):
            if len(tokens) < window:
                continue

            middle = max(0, (len(tokens) // 2) - (window // 2))
            ending = max(0, len(tokens) - window)
            if pass_index == 1:
                offsets = [0]
            else:
                offsets = [middle, ending]

            checked = set()
            for offset in offsets:
                if offset in checked:
                    continue
                checked.add(offset)
                phrase = " ".join(tokens[offset : offset + window]).strip()
                if len(phrase) < 18:
                    continue
                hits = [
                    page
                    for page in range(start_page, end_page + 1)
                    if phrase in page_texts[page - 1]
                ]
                if len(hits) == 1:
                    return hits[0], f"exact-{window}"
    return None, "none"


def weighted_fuzzy_score(needle: str, haystack: str) -> float:
    partial = float(fuzz.partial_ratio(needle, haystack))
    token_set = float(fuzz.token_set_ratio(needle, haystack))
    return (partial * 0.7) + (token_set * 0.3)


def fuzzy_page_match(
    should_norm: str,
    page_texts: List[str],
    page_range: Tuple[int, int],
) -> Tuple[int, float, Tuple[Tuple[int, float], ...], str]:
    start_page, end_page = page_range
    best_page = start_page
    best_score = -1.0
    best_method = "fuzzy"

    per_page_best: Dict[int, float] = {}

    for page in range(start_page, end_page + 1):
        current = weighted_fuzzy_score(should_norm, page_texts[page - 1])
        per_page_best[page] = max(per_page_best.get(page, -1.0), current)
        if current > best_score:
            best_score = current
            best_page = page
            best_method = "fuzzy"

        if page < end_page:
            combined = f"{page_texts[page - 1]} {page_texts[page]}"
            combined_score = weighted_fuzzy_score(should_norm, combined)
            # Keep page as starting page when best phrase spans two pages.
            per_page_best[page] = max(per_page_best.get(page, -1.0), combined_score)
            if combined_score > best_score:
                best_score = combined_score
                best_page = page
                best_method = "fuzzy-2page"

    top = sorted(per_page_best.items(), key=lambda item: item[1], reverse=True)[:3]
    return best_page, best_score, tuple(top), best_method


def match_row_to_page(
    candidate: RowCandidate,
    page_texts: List[str],
    page_range: Tuple[int, int],
) -> RowMatch:
    # Strongest signal when full normalized "Should be" appears on exactly one page.
    full_hits = [
        page
        for page in range(page_range[0], page_range[1] + 1)
        if candidate.should_norm in page_texts[page - 1]
    ]
    if len(full_hits) == 1:
        page = full_hits[0]
        return RowMatch(
            candidate=candidate,
            page=page,
            method="exact-full",
            score=100.0,
            range_start=page_range[0],
            range_end=page_range[1],
            top_candidates=((page, 100.0),),
        )

    exact_page, exact_method = longest_anchor_page_hit(
        candidate.should_norm, page_texts, page_range
    )
    if exact_page is not None:
        return RowMatch(
            candidate=candidate,
            page=exact_page,
            method=exact_method,
            score=100.0,
            range_start=page_range[0],
            range_end=page_range[1],
            top_candidates=((exact_page, 100.0),),
        )

    page, score, top, method = fuzzy_page_match(
        candidate.should_norm, page_texts, page_range
    )
    return RowMatch(
        candidate=candidate,
        page=page,
        method=method,
        score=score,
        range_start=page_range[0],
        range_end=page_range[1],
        top_candidates=top,
    )


def create_backup(path: Path) -> Path:
    stamp = dt.datetime.now().strftime("%Y%m%d-%H%M%S")
    backup = path.with_name(f"{path.stem}.bak-{stamp}{path.suffix}")
    shutil.copy2(path, backup)
    return backup


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Populate missing CRX page numbers from a print PDF."
    )
    parser.add_argument("--xlsx", required=True, help="Path to CRX workbook (.xlsx).")
    parser.add_argument("--pdf", required=True, help="Path to print PDF.")
    parser.add_argument(
        "--book-index",
        help="Optional path to book-index.json for chapter-title source.",
    )
    parser.add_argument(
        "--sheet",
        help="Worksheet name. Defaults to active worksheet.",
    )
    parser.add_argument("--start-row", type=int, default=10, help="First row to scan.")
    parser.add_argument(
        "--error-col", default="B", help="Error number column letter (default: B)."
    )
    parser.add_argument(
        "--chapter-col", default="D", help="Chapter label column letter (default: D)."
    )
    parser.add_argument(
        "--page-col", default="E", help="PDF page column letter (default: E)."
    )
    parser.add_argument(
        "--comments-col",
        default="H",
        help='Comments column letter containing "Should be:" text (default: H).',
    )
    parser.add_argument(
        "--min-confidence",
        type=float,
        default=70.0,
        help="Warn when non-exact row score is below this threshold.",
    )
    parser.add_argument(
        "--chapter-min-confidence",
        type=float,
        default=85.0,
        help="Minimum fuzzy confidence for chapter heading detection fallback.",
    )
    parser.add_argument(
        "--force",
        action="store_true",
        help="Also re-evaluate rows where page column is already populated.",
    )
    parser.add_argument(
        "--skip-low-confidence",
        action="store_true",
        help="Do not write non-exact matches below --min-confidence.",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Compute and print matches without writing workbook.",
    )
    parser.add_argument(
        "--no-backup",
        action="store_true",
        help="Do not create timestamped backup before saving.",
    )
    parser.add_argument(
        "--quiet",
        action="store_true",
        help="Reduce per-row output.",
    )
    return parser.parse_args()


def main() -> int:
    args = parse_args()

    xlsx_path = Path(args.xlsx).expanduser().resolve()
    pdf_path = Path(args.pdf).expanduser().resolve()
    book_index_path = (
        Path(args.book_index).expanduser().resolve() if args.book_index else None
    )

    if not xlsx_path.exists():
        raise FileNotFoundError(f"Workbook not found: {xlsx_path}")
    if not pdf_path.exists():
        raise FileNotFoundError(f"PDF not found: {pdf_path}")
    if book_index_path and not book_index_path.exists():
        raise FileNotFoundError(f"Book index not found: {book_index_path}")

    wb = load_workbook(xlsx_path)
    sheet = wb[args.sheet] if args.sheet else wb.active

    error_col = column_index_from_string(args.error_col)
    chapter_col = column_index_from_string(args.chapter_col)
    page_col = column_index_from_string(args.page_col)
    comments_col = column_index_from_string(args.comments_col)

    candidates = find_row_candidates(
        sheet,
        start_row=args.start_row,
        error_col=error_col,
        chapter_col=chapter_col,
        page_col=page_col,
        comments_col=comments_col,
        force=args.force,
    )
    if not candidates:
        print("No candidate rows found. Nothing to do.")
        return 0

    page_texts = read_pdf_pages(pdf_path)
    print(f"PDF pages: {len(page_texts)}")
    print(f"Candidate rows: {len(candidates)}")

    if book_index_path:
        chapter_titles = load_book_index_chapters(book_index_path)
    else:
        labels = {candidate.chapter_label for candidate in candidates}
        chapter_titles = collect_sheet_chapters(labels)

    if not chapter_titles:
        raise RuntimeError(
            "No chapter titles found. Provide --book-index or ensure column D uses 'Chapter N: Title'."
        )

    chapter_starts = detect_chapter_starts(
        chapter_titles, page_texts, chapter_min_confidence=args.chapter_min_confidence
    )
    if not chapter_starts:
        raise RuntimeError("Unable to detect chapter headings in PDF.")

    chapter_ranges = build_chapter_ranges(chapter_starts, len(page_texts))
    low_confidence: List[RowMatch] = []
    matches: List[RowMatch] = []

    for candidate in candidates:
        page_range = chapter_ranges.get(candidate.chapter_number, (1, len(page_texts)))
        match = match_row_to_page(candidate, page_texts, page_range)
        matches.append(match)
        if (not match.is_exact) and match.score < args.min_confidence:
            low_confidence.append(match)

    matches.sort(key=lambda item: item.candidate.row)

    if not args.quiet:
        for match in matches:
            top = ", ".join(f"p{page}:{score:.1f}" for page, score in match.top_candidates)
            print(
                f"row {match.candidate.row} err {match.candidate.error_id} "
                f"-> page {match.page} ({match.method}, score={match.score:.1f}, "
                f"range={match.range_start}-{match.range_end}, top={top})"
            )

    skipped = 0
    written = 0
    for match in matches:
        if (
            args.skip_low_confidence
            and (not match.is_exact)
            and match.score < args.min_confidence
        ):
            skipped += 1
            continue

        if not args.dry_run:
            sheet.cell(row=match.candidate.row, column=page_col).value = match.page
        written += 1

    backup_path: Path | None = None
    if not args.dry_run:
        if not args.no_backup:
            backup_path = create_backup(xlsx_path)
        wb.save(xlsx_path)

    print("")
    print(f"Matched rows: {len(matches)}")
    print(f"Written rows: {written}")
    print(f"Skipped rows: {skipped}")
    print(f"Low-confidence rows (< {args.min_confidence:.1f}): {len(low_confidence)}")

    if low_confidence:
        print("Low-confidence details:")
        for match in low_confidence:
            print(
                f"  row {match.candidate.row} err {match.candidate.error_id}: "
                f"page {match.page}, method={match.method}, score={match.score:.1f}"
            )

    if args.dry_run:
        print("Dry run complete. Workbook unchanged.")
    else:
        if backup_path is not None:
            print(f"Backup: {backup_path}")
        print(f"Workbook updated: {xlsx_path}")

    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except Exception as exc:
        print(f"Error: {exc}", file=sys.stderr)
        raise SystemExit(1)
