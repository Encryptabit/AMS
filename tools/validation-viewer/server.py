#!/usr/bin/env python3
"""
Validation Report Viewer Server
Serves validation reports from chapter directories with a web UI
"""

import os
import sys
import json
import re
import shutil
import subprocess
import tempfile
import traceback
import copy
from pathlib import Path
from datetime import datetime, timedelta
from http.server import HTTPServer, BaseHTTPRequestHandler
from urllib.parse import parse_qs, urlparse
from xml.etree import ElementTree as ET
from zipfile import ZipFile
import openpyxl
from openpyxl.utils import get_column_letter
import xlwings as xw

# Get AppData directory for reviewed status
APPDATA_DIR = Path(os.getenv('APPDATA')) / 'AMS' / 'validation-viewer'
APPDATA_DIR.mkdir(parents=True, exist_ok=True)
REVIEWED_STATUS_FILE = APPDATA_DIR / 'reviewed-status.json'

# Default configuration
BASE_DIR = Path(r"C:\Aethon\InProgress\Vain Glory 2")
PORT = 8081

SCRIPT_DIR = Path(__file__).resolve().parent
REPO_ROOT = SCRIPT_DIR.parent.parent
CRX_TEMPLATE_PATH = Path(r"C:\Aethon\BASE_CRX.xlsx")
CRX_DIR_NAME = 'CRX'
CRX_DATA_ROW_START = 11
DEFAULT_ERROR_TYPE = 'MR'

EXCEL_NS = 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'
XML_NS = 'http://www.w3.org/XML/1998/namespace'
REL_NS = 'http://schemas.openxmlformats.org/officeDocument/2006/relationships'
MC_NS = 'http://schemas.openxmlformats.org/markup-compatibility/2006'
X14AC_NS = 'http://schemas.microsoft.com/office/spreadsheetml/2009/9/ac'
XR_NS = 'http://schemas.microsoft.com/office/spreadsheetml/2014/revision'
XR2_NS = 'http://schemas.openxmlformats.org/spreadsheetml/2015/revision2'
XR3_NS = 'http://schemas.microsoft.com/office/spreadsheetml/2016/revision3'

NAMESPACES = {'a': EXCEL_NS}
CELL_REF_RE = re.compile(r'([A-Z]+)(\d+)')

for prefix, uri in [('', EXCEL_NS), ('r', REL_NS), ('mc', MC_NS), ('x14ac', X14AC_NS), ('xr', XR_NS), ('xr2', XR2_NS), ('xr3', XR3_NS)]:
    ET.register_namespace(prefix, uri)


class ValidationReportHandler(BaseHTTPRequestHandler):
    @staticmethod
    def load_reviewed_status():
        """Load reviewed status from AppData"""
        if REVIEWED_STATUS_FILE.exists():
            try:
                with open(REVIEWED_STATUS_FILE, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                    # Key by book name
                    book_name = BASE_DIR.name
                    return data.get(book_name, {})
            except Exception as e:
                print(f"Error loading reviewed status: {e}")
        return {}

    @staticmethod
    def save_reviewed_status(reviewed_chapters):
        """Save reviewed status to AppData"""
        try:
            # Load existing data
            all_data = {}
            if REVIEWED_STATUS_FILE.exists():
                with open(REVIEWED_STATUS_FILE, 'r', encoding='utf-8') as f:
                    all_data = json.load(f)

            # Update for current book
            book_name = BASE_DIR.name
            all_data[book_name] = reviewed_chapters

            # Save back
            with open(REVIEWED_STATUS_FILE, 'w', encoding='utf-8') as f:
                json.dump(all_data, f, indent=2)
        except Exception as e:
            print(f"Error saving reviewed status: {e}")

    def do_GET(self):
        parsed_path = urlparse(self.path)
        path = parsed_path.path

        if path == '/':
            self.serve_index()
        elif path == '/api/chapters':
            self.serve_chapters_list()
        elif path == '/api/overview':
            self.serve_overview()
        elif path == '/api/reviewed':
            self.serve_reviewed_status()
        elif path.startswith('/api/report/'):
            chapter_name = path[len('/api/report/'):]
            self.serve_report(chapter_name)
        elif path.startswith('/api/audio/'):
            # Extract chapter name and timing from path
            remainder = path[len('/api/audio/'):]
            self.serve_audio_segment(remainder, parsed_path.query)
        else:
            self.send_error(404)

    def do_POST(self):
        parsed_path = urlparse(self.path)
        path = parsed_path.path
        print(f"POST request path: {path}")

        if path.startswith('/api/export/'):
            chapter_name = path[len('/api/export/'):]
            print(f"Matched export route, chapter: {chapter_name}")
            self.handle_export_audio(chapter_name)
        elif path.startswith('/api/crx/'):
            chapter_name = path[len('/api/crx/'):]
            print(f"Matched CRX route, chapter: {chapter_name}")
            self.handle_add_to_crx(chapter_name)
        elif path.startswith('/api/reviewed/'):
            chapter_name = path[len('/api/reviewed/'):]
            print(f"Matched reviewed route, chapter: {chapter_name}")
            self.handle_mark_reviewed(chapter_name)
        elif path == '/api/reset-reviews':
            print(f"Matched reset reviews route")
            self.handle_reset_reviews()
        else:
            print(f"No route matched for path: {path}")
            self.send_error(404)

    def serve_reviewed_status(self):
        """Serve reviewed status for all chapters"""
        reviewed = self.load_reviewed_status()
        self.send_json_response(reviewed)

    def handle_mark_reviewed(self, chapter_name):
        """Mark a chapter as reviewed"""
        from urllib.parse import unquote

        chapter_name = unquote(chapter_name)

        # Read POST data
        content_length = int(self.headers['Content-Length'])
        post_data = self.rfile.read(content_length)
        params = json.loads(post_data.decode('utf-8'))

        reviewed_status = params.get('reviewed', True)

        try:
            # Load current status
            all_reviewed = self.load_reviewed_status()

            # Update status for this chapter
            all_reviewed[chapter_name] = {
                'reviewed': reviewed_status,
                'timestamp': datetime.utcnow().isoformat()
            }

            # Save back
            self.save_reviewed_status(all_reviewed)

            self.send_json_response({
                'success': True,
                'chapter': chapter_name,
                'reviewed': reviewed_status
            })

        except Exception as e:
            print(f"Error marking reviewed: {e}")
            print(traceback.format_exc())
            self.send_json_response({'error': str(e)}, 500)

    def handle_reset_reviews(self):
        """Reset all review status for the current book"""
        try:
            # Clear all reviewed status for current book
            self.save_reviewed_status({})

            self.send_json_response({
                'success': True,
                'message': 'All review status reset'
            })

        except Exception as e:
            print(f"Error resetting reviews: {e}")
            print(traceback.format_exc())
            self.send_json_response({'error': str(e)}, 500)

    def serve_index(self):
        """Serve the main HTML page"""
        html = """<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Validation Report Viewer</title>
    <style>
        * {
            margin: 0;
            padding: 0;
            box-sizing: border-box;
        }

        body {
            font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, Oxygen, Ubuntu, Cantarell, sans-serif;
            background: #1e1e1e;
            color: #d4d4d4;
            height: 100vh;
            display: flex;
        }

        #sidebar {
            width: 280px;
            background: #252526;
            border-right: 1px solid #3e3e42;
            overflow-y: auto;
            flex-shrink: 0;
        }

        #sidebar h2 {
            padding: 20px;
            font-size: 16px;
            font-weight: 600;
            color: #cccccc;
            border-bottom: 1px solid #3e3e42;
            background: #2d2d30;
        }

        .chapter-item {
            padding: 12px 20px;
            cursor: pointer;
            border-bottom: 1px solid #3e3e42;
            transition: background 0.2s;
        }

        .chapter-item:hover {
            background: #2a2d2e;
        }

        .chapter-item.active {
            background: #094771;
            border-left: 3px solid #007acc;
        }

        .chapter-name {
            font-size: 14px;
            font-weight: 500;
        }

        .chapter-stats {
            font-size: 12px;
            color: #858585;
            margin-top: 4px;
        }

        #content {
            flex: 1;
            overflow-y: auto;
            padding: 30px;
        }

        #loading {
            text-align: center;
            padding: 40px;
            color: #858585;
        }

        .report-header {
            background: #252526;
            padding: 24px;
            border-radius: 6px;
            margin-bottom: 24px;
            border: 1px solid #3e3e42;
        }

        .report-header h1 {
            font-size: 24px;
            margin-bottom: 16px;
            color: #ffffff;
        }

        .report-meta {
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(200px, 1fr));
            gap: 12px;
            font-size: 13px;
        }

        .meta-item {
            color: #cccccc;
        }

        .meta-label {
            color: #858585;
            font-weight: 500;
        }

        .stats-grid {
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(250px, 1fr));
            gap: 16px;
            margin-bottom: 32px;
        }

        .stat-card {
            background: #252526;
            padding: 20px;
            border-radius: 6px;
            border: 1px solid #3e3e42;
        }

        .stat-label {
            font-size: 12px;
            color: #858585;
            text-transform: uppercase;
            font-weight: 600;
            letter-spacing: 0.5px;
            margin-bottom: 8px;
        }

        .stat-value {
            font-size: 28px;
            font-weight: 600;
            color: #ffffff;
        }

        .stat-detail {
            font-size: 13px;
            color: #cccccc;
            margin-top: 8px;
        }

        .section-title {
            font-size: 18px;
            font-weight: 600;
            margin: 32px 0 16px;
            color: #ffffff;
            border-bottom: 2px solid #3e3e42;
            padding-bottom: 8px;
        }

        .sentence-item {
            background: #252526;
            padding: 20px;
            border-radius: 6px;
            margin-bottom: 16px;
            border-left: 4px solid #3e3e42;
        }

        .sentence-item.unreliable {
            border-left-color: #f48771;
        }

        .sentence-item.attention {
            border-left-color: #dcdcaa;
        }

        .sentence-item.highlight {
            box-shadow: 0 0 0 2px rgba(255, 215, 0, 0.25);
            border-left-color: #ffd54f;
        }

        .sentence-header {
            display: flex;
            justify-content: space-between;
            align-items: center;
            margin-bottom: 12px;
            flex-wrap: wrap;
            gap: 8px;
        }

        .sentence-id {
            font-size: 14px;
            font-weight: 600;
            color: #4ec9b0;
        }

        .sentence-metrics {
            display: flex;
            gap: 16px;
            font-size: 13px;
        }

        .metric {
            display: flex;
            align-items: center;
            gap: 6px;
        }

        .metric-label {
            color: #858585;
        }

        .metric-value {
            font-weight: 600;
        }

        .metric-value.high {
            color: #f48771;
        }

        .metric-value.medium {
            color: #dcdcaa;
        }

        .metric-value.low {
            color: #4ec9b0;
        }

        .status-badge {
            padding: 4px 10px;
            border-radius: 4px;
            font-size: 11px;
            font-weight: 600;
            text-transform: uppercase;
            letter-spacing: 0.5px;
        }

        .status-badge.unreliable {
            background: #f487711a;
            color: #f48771;
        }

        .status-badge.attention {
            background: #dcdcaa1a;
            color: #dcdcaa;
        }

        .sentence-details {
            display: grid;
            gap: 8px;
            font-size: 13px;
            color: #cccccc;
            margin-bottom: 12px;
        }

        .paragraph-flags {
            font-size: 13px;
            margin-top: 4px;
        }

        .paragraph-flags a {
            color: #4ec9b0;
            text-decoration: none;
        }

        .paragraph-flags a:hover {
            text-decoration: underline;
        }

        .sentence-text {
            background: #1e1e1e;
            padding: 12px;
            border-radius: 4px;
            font-family: 'Consolas', 'Courier New', monospace;
            line-height: 1.6;
            margin-top: 8px;
        }

        .text-label {
            font-size: 11px;
            color: #858585;
            text-transform: uppercase;
            font-weight: 600;
            margin-bottom: 4px;
        }

        .text-content {
            color: #d4d4d4;
        }

        .diff-highlight {
            background: #3a3a00;
            color: #ffff99;
            padding: 2px 4px;
            border-radius: 3px;
            font-weight: 500;
        }

        /* Inline diff view styles */
        .diff-unified {
            background: #1e1e1e;
            border: 1px solid #3e3e42;
            border-radius: 4px;
            padding: 12px;
        }

        .diff-inline {
            color: #d4d4d4;
            line-height: 1.8;
            font-size: 14px;
            word-wrap: break-word;
        }

        .diff-deleted {
            background-color: rgba(255, 0, 0, 0.2);
            color: #ff6b6b;
            text-decoration: line-through;
            padding: 2px 4px;
            border-radius: 3px;
            margin: 0 1px;
        }

        .diff-inserted {
            background-color: rgba(0, 255, 0, 0.2);
            color: #69db7c;
            padding: 2px 4px;
            border-radius: 3px;
            margin: 0 1px;
            font-weight: 500;
        }

        .diff-empty {
            color: #888888;
            font-style: italic;
            padding: 8px;
        }

        .audio-player {
            margin-top: 12px;
            padding: 12px;
            background: #1e1e1e;
            border-radius: 4px;
            border: 1px solid #3e3e42;
        }

        .audio-player audio {
            width: 100%;
            height: 32px;
        }

        .play-button {
            display: inline-flex;
            align-items: center;
            gap: 6px;
            padding: 6px 12px;
            background: #0e639c;
            color: #ffffff;
            border: none;
            border-radius: 4px;
            cursor: pointer;
            font-size: 12px;
            font-weight: 500;
            transition: background 0.2s;
        }

        .play-button:hover {
            background: #1177bb;
        }

        .play-button:active {
            background: #0d5a8f;
        }

        .play-icon {
            width: 0;
            height: 0;
            border-left: 8px solid currentColor;
            border-top: 5px solid transparent;
            border-bottom: 5px solid transparent;
        }

        .action-buttons {
            display: flex;
            gap: 8px;
            margin-top: 12px;
        }

        .export-button {
            display: inline-flex;
            align-items: center;
            gap: 6px;
            padding: 6px 12px;
            background: #0e7a0d;
            color: #ffffff;
            border: none;
            border-radius: 4px;
            cursor: pointer;
            font-size: 12px;
            font-weight: 500;
            transition: background 0.2s;
        }

        .export-button:hover {
            background: #0d9e0c;
        }

        .export-button:active {
            background: #0a660a;
        }

        .crx-button {
            display: inline-flex;
            align-items: center;
            gap: 6px;
            padding: 6px 12px;
            background: #7a0e7a;
            color: #ffffff;
            border: none;
            border-radius: 4px;
            cursor: pointer;
            font-size: 12px;
            font-weight: 500;
            transition: background 0.2s;
        }

        .crx-button:hover {
            background: #9e0c9e;
        }

        .crx-button:active {
            background: #660a66;
        }

        .button-disabled {
            opacity: 0.5;
            cursor: not-allowed;
        }

        .modal {
            display: none;
            position: fixed;
            z-index: 1000;
            left: 0;
            top: 0;
            width: 100%;
            height: 100%;
            background-color: rgba(0, 0, 0, 0.6);
        }

        .modal-content {
            background-color: #252526;
            margin: 15% auto;
            padding: 24px;
            border: 1px solid #3e3e42;
            border-radius: 8px;
            width: 80%;
            max-width: 500px;
        }

        .modal-header {
            font-size: 18px;
            font-weight: 600;
            margin-bottom: 16px;
            color: #ffffff;
        }

        .modal-body {
            margin-bottom: 16px;
        }

        .modal-input {
            width: 100%;
            padding: 8px;
            margin: 8px 0;
            background: #1e1e1e;
            color: #d4d4d4;
            border: 1px solid #3e3e42;
            border-radius: 4px;
            font-size: 13px;
        }

        .modal-textarea {
            width: 100%;
            padding: 8px;
            margin: 8px 0;
            background: #1e1e1e;
            color: #d4d4d4;
            border: 1px solid #3e3e42;
            border-radius: 4px;
            font-size: 13px;
            min-height: 80px;
            resize: vertical;
        }

        .modal-footer {
            display: flex;
            justify-content: flex-end;
            gap: 8px;
        }

        .modal-button {
            padding: 8px 16px;
            border: none;
            border-radius: 4px;
            cursor: pointer;
            font-size: 13px;
            font-weight: 500;
        }

        .modal-button-primary {
            background: #0e639c;
            color: #ffffff;
        }

        .modal-button-primary:hover {
            background: #1177bb;
        }

        .modal-button-secondary {
            background: #3e3e42;
            color: #d4d4d4;
        }

        .modal-button-secondary:hover {
            background: #4e4e52;
        }

        .paragraph-item {
            background: #252526;
            padding: 18px;
            border-radius: 6px;
            margin-bottom: 12px;
            border-left: 3px solid #3e3e42;
        }

        .paragraph-item.unreliable {
            border-left-color: #f48771;
        }

        .paragraph-item.attention {
            border-left-color: #dcdcaa;
        }

        ::-webkit-scrollbar {
            width: 10px;
            height: 10px;
        }

        ::-webkit-scrollbar-track {
            background: #1e1e1e;
        }

        ::-webkit-scrollbar-thumb {
            background: #424242;
            border-radius: 5px;
        }

        ::-webkit-scrollbar-thumb:hover {
            background: #4e4e4e;
        }

        .toast-container {
            position: fixed;
            bottom: 20px;
            right: 20px;
            z-index: 2000;
            display: flex;
            flex-direction: column;
            gap: 10px;
        }

        .toast {
            background: #252526;
            border: 1px solid #3e3e42;
            border-radius: 6px;
            padding: 16px 20px;
            min-width: 300px;
            max-width: 400px;
            box-shadow: 0 4px 12px rgba(0, 0, 0, 0.5);
            animation: slideIn 0.3s ease-out;
        }

        @keyframes slideIn {
            from {
                transform: translateX(400px);
                opacity: 0;
            }
            to {
                transform: translateX(0);
                opacity: 1;
            }
        }

        .toast.success {
            border-left: 4px solid #4ec9b0;
        }

        .toast.error {
            border-left: 4px solid #f48771;
        }

        .toast-header {
            display: flex;
            justify-content: space-between;
            align-items: center;
            margin-bottom: 8px;
        }

        .toast-title {
            font-weight: 600;
            font-size: 14px;
            color: #ffffff;
        }

        .toast-close {
            background: none;
            border: none;
            color: #858585;
            cursor: pointer;
            font-size: 18px;
            padding: 0;
            line-height: 1;
        }

        .toast-close:hover {
            color: #d4d4d4;
        }

        .toast-message {
            font-size: 13px;
            color: #cccccc;
            line-height: 1.5;
        }

        .chapter-grid {
            display: grid;
            grid-template-columns: repeat(auto-fill, minmax(350px, 1fr));
            gap: 16px;
            margin-top: 16px;
        }

        .chapter-card {
            background: #252526;
            border: 1px solid #3e3e42;
            border-radius: 6px;
            padding: 16px;
            cursor: pointer;
            transition: all 0.2s;
        }

        .chapter-card:hover {
            background: #2d2d30;
            border-color: #007acc;
            box-shadow: 0 2px 8px rgba(0, 122, 204, 0.2);
        }

        .chapter-card-header {
            margin-bottom: 12px;
            padding-bottom: 12px;
            border-bottom: 1px solid #3e3e42;
        }

        .chapter-card-name {
            font-size: 15px;
            font-weight: 600;
            color: #ffffff;
        }

        .chapter-card-metrics {
            display: flex;
            flex-direction: column;
            gap: 8px;
            font-size: 13px;
        }

        .chapter-card-metrics .metric {
            display: flex;
            justify-content: space-between;
        }

        .overview-item {
            border-left: 3px solid #007acc;
            background: #2a2d2e;
        }

        .reviewed-button {
            position: fixed;
            bottom: 30px;
            right: 30px;
            display: inline-flex;
            align-items: center;
            gap: 8px;
            padding: 12px 24px;
            background: #0e7a0d;
            color: #ffffff;
            border: none;
            border-radius: 6px;
            cursor: pointer;
            font-size: 14px;
            font-weight: 600;
            transition: all 0.2s;
            box-shadow: 0 4px 12px rgba(0, 0, 0, 0.5);
            z-index: 1000;
        }

        .reviewed-button:hover {
            background: #0d9e0c;
            transform: translateY(-2px);
            box-shadow: 0 6px 16px rgba(0, 0, 0, 0.6);
        }

        .reviewed-button.reviewed {
            background: #3e3e42;
            color: #4ec9b0;
        }

        .reviewed-button.reviewed:hover {
            background: #4e4e52;
            transform: translateY(-2px);
            box-shadow: 0 6px 16px rgba(0, 0, 0, 0.6);
        }

        .reset-review-button {
            padding: 8px 16px;
            background: #722020;
            color: #ffffff;
            border: none;
            border-radius: 4px;
            cursor: pointer;
            font-size: 13px;
            font-weight: 500;
            transition: all 0.2s;
        }

        .reset-review-button:hover {
            background: #8b2828;
            transform: translateY(-1px);
        }

        .reviewed-indicator {
            display: inline-flex;
            align-items: center;
            gap: 4px;
            padding: 2px 8px;
            background: #4ec9b01a;
            color: #4ec9b0;
            border-radius: 3px;
            font-size: 11px;
            font-weight: 600;
            margin-left: 8px;
        }

        .chapter-card.reviewed {
            opacity: 0.7;
        }

        .chapter-card.reviewed .chapter-card-name::after {
            content: ' âœ“';
            color: #4ec9b0;
            margin-left: 8px;
        }

        .chapter-item.reviewed {
            opacity: 0.7;
        }

        .chapter-item.reviewed .chapter-name::after {
            content: ' âœ“';
            color: #4ec9b0;
            margin-left: 4px;
        }
    </style>
</head>
<body>
    <div id="toast-container" class="toast-container"></div>
    <div id="sidebar">
        <h2>Validation Reports</h2>
        <div id="chapter-list"></div>
    </div>
    <div id="content">
        <div id="loading">Select a chapter to view its validation report</div>
    </div>

    <!-- CRX Modal -->
    <div id="crxModal" class="modal">
        <div class="modal-content">
            <div class="modal-header">Add to CRX</div>
            <div class="modal-body">
                <label for="errorType">Error Type:</label>
                <select id="errorType" class="modal-input">
                    <option value="MR" selected>MR</option>
                    <option value="PRON">PRON</option>
                    <option value="DIC">DIC</option>
                    <option value="NZ">NZ</option>
                    <option value="PL">PL</option>
                    <option value="DIST">DIST</option>
                    <option value="MW">MW</option>
                    <option value="ML">ML</option>
                    <option value="TYPO">TYPO</option>
                    <option value="CHAR">CHAR</option>
                </select>
                <label for="comments">Comments:</label>
                <textarea id="comments" class="modal-textarea" placeholder="Enter any additional comments..."></textarea>
            </div>
            <div class="modal-footer">
                <button class="modal-button modal-button-secondary" onclick="closeCrxModal()">Cancel</button>
                <button class="modal-button modal-button-primary" onclick="submitCrx()">Add to CRX</button>
            </div>
        </div>
    </div>

    <script>
        let chapters = [];
        let currentChapter = null;
        let currentReport = null;
        let pendingCrxData = null;
        let reviewedStatus = {};

        async function loadReviewedStatus() {
            try {
                const response = await fetch('/api/reviewed');
                reviewedStatus = await response.json();
            } catch (error) {
                console.error('Failed to load reviewed status:', error);
                reviewedStatus = {};
            }
        }

        async function loadChapters() {
            try {
                const response = await fetch('/api/chapters');
                chapters = await response.json();
                await loadReviewedStatus();
                renderChapterList();
            } catch (error) {
                console.error('Failed to load chapters:', error);
            }
        }

        async function markReviewed(chapterName, reviewed = true) {
            try {
                const response = await fetch(`/api/reviewed/${encodeURIComponent(chapterName)}`, {
                    method: 'POST',
                    headers: {'Content-Type': 'application/json'},
                    body: JSON.stringify({ reviewed: reviewed })
                });

                const result = await response.json();

                if (result.success) {
                    reviewedStatus[chapterName] = { reviewed: reviewed, timestamp: new Date().toISOString() };
                    showToast(reviewed ? `Marked "${chapterName}" as reviewed` : `Unmarked "${chapterName}"`, 'success');
                    renderChapterList();
                    // Re-render current view
                    if (currentChapter === chapterName) {
                        loadReport(chapterName);
                    } else if (!currentChapter) {
                        loadOverview();
                    }
                    // Update floating button
                    updateFloatingButton();
                } else {
                    showToast(`Failed to update reviewed status: ${result.error}`, 'error');
                }
            } catch (err) {
                console.error('Failed to mark reviewed:', err);
                showToast('Failed to update reviewed status', 'error');
            }
        }

        async function resetAllReviews() {
            if (!confirm('Are you sure you want to reset review status for ALL chapters? This cannot be undone.')) {
                return;
            }

            try {
                const response = await fetch('/api/reset-reviews', {
                    method: 'POST',
                    headers: { 'Content-Type': 'application/json' }
                });

                const result = await response.json();

                if (result.success) {
                    reviewedStatus = {};
                    showToast('Reset all review status', 'success');
                    renderChapterList();
                    loadOverview(); // Reload overview to update visual state
                } else {
                    showToast(`Failed to reset reviews: ${result.error}`, 'error');
                }
            } catch (err) {
                console.error('Failed to reset reviews:', err);
                showToast('Failed to reset review status', 'error');
            }
        }

        function updateFloatingButton() {
            // Remove existing button
            const existingButton = document.getElementById('floating-reviewed-button');
            if (existingButton) {
                existingButton.remove();
            }

            // Only show button when viewing a chapter report
            if (!currentChapter) return;

            const isReviewed = reviewedStatus[currentChapter]?.reviewed;
            const reviewedClass = isReviewed ? 'reviewed' : '';
            const buttonText = isReviewed ? 'âœ“ Reviewed' : 'Mark as Reviewed';

            const button = document.createElement('button');
            button.id = 'floating-reviewed-button';
            button.className = `reviewed-button ${reviewedClass}`;
            button.innerHTML = buttonText;
            button.onclick = () => markReviewed(currentChapter, !isReviewed);

            document.body.appendChild(button);
        }

        function renderChapterList() {
            const list = document.getElementById('chapter-list');

            // Add overview link at the top
            let html = `
                <div class="chapter-item overview-item" onclick="loadOverview()">
                    <div class="chapter-name">ðŸ“Š Book Overview</div>
                    <div class="chapter-stats">All Chapters</div>
                </div>
            `;

            html += chapters.map(chapter => {
                const m = chapter.metrics;
                const isReviewed = reviewedStatus[chapter.name]?.reviewed;
                const reviewedClass = isReviewed ? 'reviewed' : '';
                const flaggedInfo = `${m.sentenceFlagged} sentences, ${m.paragraphFlagged} paragraphs`;
                return `
                    <div class="chapter-item ${reviewedClass}" onclick="loadReport('${chapter.name}')">
                        <div class="chapter-name">${chapter.name}</div>
                        <div class="chapter-stats">Flagged: ${flaggedInfo}</div>
                        <div class="chapter-stats">Avg WER: ${m.sentenceAvgWer}</div>
                    </div>
                `;
            }).join('');

            list.innerHTML = html;
        }

        async function loadOverview() {
            currentChapter = null;

            // Update active state
            document.querySelectorAll('.chapter-item').forEach(item => {
                if (item.classList.contains('overview-item')) {
                    item.classList.add('active');
                } else {
                    item.classList.remove('active');
                }
            });

            try {
                const response = await fetch('/api/overview');
                const overview = await response.json();
                renderOverview(overview);
                updateFloatingButton(); // Hide button on overview
                // Scroll the content div to top (not window, since content has overflow-y: auto)
                const contentDiv = document.getElementById('content');
                if (contentDiv) {
                    contentDiv.scrollTo({ top: 0, behavior: 'smooth' });
                }
            } catch (error) {
                console.error('Failed to load overview:', error);
                document.getElementById('content').innerHTML = '<div id="loading">Failed to load overview</div>';
            }
        }

        function renderOverview(overview) {
            const content = document.getElementById('content');

            const chaptersHtml = overview.chapters.map(c => {
                const m = c.metrics;
                const werClass = parseFloat(m.sentenceAvgWer) > 5 ? 'high' : parseFloat(m.sentenceAvgWer) > 2 ? 'medium' : 'low';
                const isReviewed = reviewedStatus[c.name]?.reviewed;
                const reviewedClass = isReviewed ? 'reviewed' : '';

                return `
                    <div class="chapter-card ${reviewedClass}" onclick="loadReport('${c.name}')">
                        <div class="chapter-card-header">
                            <div class="chapter-card-name">${c.name}</div>
                        </div>
                        <div class="chapter-card-metrics">
                            <div class="metric">
                                <span class="metric-label">Sentences:</span>
                                <span class="metric-value">${m.sentenceCount} (${m.sentenceFlagged} flagged)</span>
                            </div>
                            <div class="metric">
                                <span class="metric-label">Paragraphs:</span>
                                <span class="metric-value">${m.paragraphCount} (${m.paragraphFlagged} flagged)</span>
                            </div>
                            <div class="metric">
                                <span class="metric-label">Avg WER:</span>
                                <span class="metric-value ${werClass}">${m.sentenceAvgWer}</span>
                            </div>
                        </div>
                    </div>
                `;
            }).join('');

            content.innerHTML = `
                <div class="report-header">
                    <h1>${overview.bookName} - Validation Overview</h1>
                </div>

                <div class="stats-grid">
                    <div class="stat-card">
                        <div class="stat-label">Total Chapters</div>
                        <div class="stat-value">${overview.chapterCount}</div>
                    </div>
                    <div class="stat-card">
                        <div class="stat-label">Total Sentences</div>
                        <div class="stat-value">${overview.totalSentences}</div>
                        <div class="stat-detail">Flagged: ${overview.totalFlaggedSentences}</div>
                        <div class="stat-detail">Avg WER: ${overview.avgSentenceWer}</div>
                    </div>
                    <div class="stat-card">
                        <div class="stat-label">Total Paragraphs</div>
                        <div class="stat-value">${overview.totalParagraphs}</div>
                        <div class="stat-detail">Flagged: ${overview.totalFlaggedParagraphs}</div>
                        <div class="stat-detail">Avg WER: ${overview.avgParagraphWer}</div>
                    </div>
                </div>

                <div class="section-title" style="display: flex; justify-content: space-between; align-items: center;">
                    <span>Chapters</span>
                    <button class="reset-review-button" onclick="resetAllReviews()">Reset All Review Status</button>
                </div>
                <div class="chapter-grid">
                    ${chaptersHtml}
                </div>
            `;
        }

        async function loadReport(chapterName) {
            currentChapter = chapterName;

            // Update active state - need to account for overview item at index 0
            document.querySelectorAll('.chapter-item').forEach((item) => {
                if (item.classList.contains('overview-item')) {
                    item.classList.remove('active');
                } else {
                    const itemName = item.querySelector('.chapter-name').textContent;
                    if (itemName === chapterName) {
                        item.classList.add('active');
                    } else {
                        item.classList.remove('active');
                    }
                }
            });

            try {
                const response = await fetch(`/api/report/${encodeURIComponent(chapterName)}`);
                const report = await response.json();
                renderReport(report);
                updateFloatingButton(); // Show/update button for this chapter
                // Scroll the content div to top (not window, since content has overflow-y: auto)
                const contentDiv = document.getElementById('content');
                if (contentDiv) {
                    contentDiv.scrollTo({ top: 0, behavior: 'smooth' });
                }
            } catch (error) {
                console.error('Failed to load report:', error);
                document.getElementById('content').innerHTML = '<div id="loading">Failed to load report</div>';
            }
        }

        function renderReport(report) {
            currentReport = report;
            const content = document.getElementById('content');

            const sentencesHtml = report.sentences.map(s => `
                <div id="sentence-${s.id}" class="sentence-item ${s.status}">
                    <div class="sentence-header">
                        <span class="sentence-id">#${s.id}</span>
                        <div class="sentence-metrics">
                            <div class="metric">
                                <span class="metric-label">WER:</span>
                                <span class="metric-value ${getMetricClass(s.wer)}">${s.wer}</span>
                            </div>
                            <div class="metric">
                                <span class="metric-label">CER:</span>
                                <span class="metric-value ${getMetricClass(s.cer)}">${s.cer}</span>
                            </div>
                            <span class="status-badge ${s.status}">${s.status}</span>
                        </div>
                    </div>
                    <div class="sentence-details">
                        <div>Book range: ${s.bookRange}</div>
                        <div>Script range: ${s.scriptRange}</div>
                        <div>Timing: ${s.timing}</div>
                        ${s.paragraphId !== null && s.paragraphId !== undefined ? `<div class="paragraph-flags">Parent paragraph: <a href="#paragraph-${s.paragraphId}" onclick="focusParagraph(${s.paragraphId}); return false;">#${s.paragraphId}</a></div>` : ''}
                    </div>
                    <div class="sentence-text">
                        ${s.diff && s.diff.ops ? `
                            <div class="text-label">Manuscript</div>
                            <div class="text-content">${escapeHtml(s.bookText)}</div>
                            <div class="text-label" style="margin-top: 12px;">Transcript</div>
                            <div class="diff-unified" data-sentence-id="${s.id}"></div>
                        ` : s.scriptText && s.bookText ? `
                            <div class="text-label">Script (Read as)</div>
                            <div class="text-content diff-text" data-sentence-id="${s.id}" data-diff-script="${escapeHtml(s.scriptText)}" data-diff-book="${escapeHtml(s.bookText)}"></div>
                            <div class="text-label" style="margin-top: 12px;">Book (Should be)</div>
                            <div class="text-content diff-text" data-sentence-id="${s.id}" data-diff-script="${escapeHtml(s.scriptText)}" data-diff-book="${escapeHtml(s.bookText)}" data-diff-type="book"></div>
                        ` : `
                            <div class="text-label">Book</div>
                            <div class="text-content">${escapeHtml(s.bookText)}</div>
                            ${s.scriptText ? `
                                <div class="text-label" style="margin-top: 12px;">Script</div>
                                <div class="text-content">${escapeHtml(s.scriptText)}</div>
                            ` : ''}
                        `}
                    </div>
                    ${s.startTime !== null && s.endTime !== null ? `
                        <div class="action-buttons">
                            <button class="play-button" onclick="playAudioSegment('${currentChapter}', ${s.startTime}, ${s.endTime})">
                                <span class="play-icon"></span>
                                Play Audio Segment
                            </button>
                            <button class="export-button" data-chapter="${currentChapter}" data-start="${s.startTime}" data-end="${s.endTime}" data-sentence-id="${s.id}" data-excerpt="${escapeHtml(s.excerpt || '')}">
                                Export Audio
                            </button>
                            <button class="crx-button" data-chapter="${currentChapter}" data-start="${s.startTime}" data-end="${s.endTime}" data-sentence-id="${s.id}" data-excerpt="${escapeHtml(s.excerpt || '')}">
                                Add to CRX
                            </button>
                        </div>
                    ` : ''}
                </div>
            `).join('');

            const paragraphsHtml = report.paragraphs.map(p => `
                <div id="paragraph-${p.id}" class="paragraph-item ${p.status}">
                    <div class="sentence-header">
                        <span class="sentence-id">#${p.id}</span>
                        <div class="sentence-metrics">
                            <div class="metric">
                                <span class="metric-label">WER:</span>
                                <span class="metric-value ${getMetricClass(parseFloat(p.wer))}">${p.wer}</span>
                            </div>
                            <div class="metric">
                                <span class="metric-label">Coverage:</span>
                                <span class="metric-value">${p.coverage}</span>
                            </div>
                            <span class="status-badge ${p.status}">${p.status}</span>
                        </div>
                    </div>
                    <div class="sentence-details">
                        <div>Book range: ${p.bookRange}</div>
                        ${p.timing ? `<div>Timing: ${p.timing}</div>` : ''}
                        ${Array.isArray(p.flaggedSentenceIds) && p.flaggedSentenceIds.length ? `<div class="paragraph-flags">Flagged sentences: ${p.flaggedSentenceIds.map(id => `<a href="#sentence-${id}" onclick="focusSentence(${id}); return false;">#${id}</a>`).join(', ')}</div>` : ''}
                    </div>
                    ${p.bookText ? `
                        <div class="sentence-text">
                            <div class="text-label">Book</div>
                            <div class="text-content">${renderParagraphText(p.bookText, p.flaggedSentenceIds || [], report.sentences)}</div>
                        </div>
                    ` : ''}
                    ${p.startTime !== null && p.endTime !== null ? `
                        <div class="action-buttons">
                            <button class="play-button" onclick="playAudioSegment('${currentChapter}', ${p.startTime}, ${p.endTime})">
                                <span class="play-icon"></span>
                                Play Paragraph Audio
                            </button>
                        </div>
                    ` : ''}
                </div>
            `).join('');

            content.innerHTML = `
                <div class="report-header">
                    <h1>${report.chapterName}</h1>
                    <div class="report-meta">
                        <div class="meta-item"><span class="meta-label">Audio:</span> ${report.audioPath}</div>
                        <div class="meta-item"><span class="meta-label">Script:</span> ${report.scriptPath}</div>
                        <div class="meta-item"><span class="meta-label">Created:</span> ${report.created}</div>
                    </div>
                </div>

                <div class="stats-grid">
                    <div class="stat-card">
                        <div class="stat-label">Sentences</div>
                        <div class="stat-value">${report.stats.sentenceCount}</div>
                        <div class="stat-detail">
                            Avg WER ${report.stats.avgWer} | Max WER ${report.stats.maxWer}
                        </div>
                        <div class="stat-detail">Flagged: ${report.stats.flaggedCount}</div>
                    </div>
                    <div class="stat-card">
                        <div class="stat-label">Paragraphs</div>
                        <div class="stat-value">${report.stats.paragraphCount}</div>
                        <div class="stat-detail">
                            Avg WER ${report.stats.paragraphAvgWer}
                        </div>
                        <div class="stat-detail">Avg Coverage ${report.stats.avgCoverage}</div>
                    </div>
                </div>

                <div class="section-title">Sentences by WER</div>
                ${sentencesHtml}

                <div class="section-title">Paragraphs by WER</div>
                ${paragraphsHtml}
            `;

            // Add event listeners for export and CRX buttons
            setTimeout(() => {
                document.querySelectorAll('.export-button').forEach(btn => {
                    btn.addEventListener('click', function() {
                        const chapter = this.getAttribute('data-chapter');
                        const start = parseFloat(this.getAttribute('data-start'));
                        const end = parseFloat(this.getAttribute('data-end'));
                        const sentenceId = this.getAttribute('data-sentence-id');
                        const excerpt = this.getAttribute('data-excerpt');
                        exportAudio(chapter, start, end, sentenceId, excerpt);
                    });
                });

                document.querySelectorAll('.crx-button').forEach(btn => {
                    btn.addEventListener('click', function() {
                        const chapter = this.getAttribute('data-chapter');
                        const start = parseFloat(this.getAttribute('data-start'));
                        const end = parseFloat(this.getAttribute('data-end'));
                        const sentenceId = parseInt(this.getAttribute('data-sentence-id'), 10);
                        const excerpt = this.getAttribute('data-excerpt');
                        openCrxModal(chapter, start, end, sentenceId, excerpt);
                    });
                });

                // Apply diff highlighting to all diff-text elements
                document.querySelectorAll('.diff-text').forEach(elem => {
                    const sentenceIdAttr = elem.getAttribute('data-sentence-id');
                    const diffType = elem.getAttribute('data-diff-type') || 'script';
                    const scriptRaw = elem.getAttribute('data-diff-script');
                    const bookRaw = elem.getAttribute('data-diff-book');

                    const scriptText = decodeHtml(scriptRaw);
                    const bookText = decodeHtml(bookRaw);

                    let wordOps = null;
                    if (sentenceIdAttr && currentReport && Array.isArray(currentReport.sentences)) {
                        const sentenceId = parseInt(sentenceIdAttr, 10);
                        const sentenceData = currentReport.sentences.find(s => s.id === sentenceId);
                        if (sentenceData && Array.isArray(sentenceData.wordOps)) {
                            wordOps = sentenceData.wordOps;
                        }
                    }

                    if (scriptText || bookText) {
                        const highlighted = highlightDifferencesVisual(scriptText, bookText, wordOps);
                        elem.innerHTML = diffType === 'book' ? highlighted.bookHighlighted : highlighted.scriptHighlighted;
                    }
                });

                // Render unified diff views
                document.querySelectorAll('.diff-unified').forEach(elem => {
                    const sentenceIdAttr = elem.getAttribute('data-sentence-id');
                    if (sentenceIdAttr && currentReport && Array.isArray(currentReport.sentences)) {
                        const sentenceId = parseInt(sentenceIdAttr, 10);
                        const sentenceData = currentReport.sentences.find(s => s.id === sentenceId);
                        if (sentenceData && sentenceData.diff) {
                            elem.innerHTML = renderUnifiedDiff(sentenceData.diff);
                        }
                    }
                });
            }, 0);
        }

        function getMetricClass(value) {
            if (value >= 50) return 'high';
            if (value >= 10) return 'medium';
            return 'low';
        }

        function escapeHtml(text) {
            const div = document.createElement('div');
            div.textContent = text;
            return div.innerHTML;
        }

        function decodeHtml(html) {
            if (!html) return '';
            const textarea = document.createElement('textarea');
            textarea.innerHTML = html;
            return textarea.value;
        }

        function playAudioSegment(chapterName, startTime, endTime) {
            // Create or reuse audio element
            let audioPlayer = document.getElementById('global-audio-player');
            if (!audioPlayer) {
                audioPlayer = document.createElement('audio');
                audioPlayer.id = 'global-audio-player';
                audioPlayer.controls = true;
                audioPlayer.style.display = 'none';
                document.body.appendChild(audioPlayer);
            }

            // Set source with timing parameters
            const url = `/api/audio/${encodeURIComponent(chapterName)}?start=${startTime}&end=${endTime}`;
            audioPlayer.src = url;
            audioPlayer.currentTime = 0;
            audioPlayer.play().catch(err => {
                console.error('Failed to play audio:', err);
                alert('Failed to play audio segment. Check console for details.');
            });
        }

        function focusSentence(sentenceId) {
            const element = document.getElementById(`sentence-${sentenceId}`);
            if (!element) return;

            element.scrollIntoView({ behavior: 'smooth', block: 'center' });
            element.classList.add('highlight');
            setTimeout(() => element.classList.remove('highlight'), 1600);
        }

        function focusParagraph(paragraphId) {
            const element = document.getElementById(`paragraph-${paragraphId}`);
            if (!element) return;

            element.scrollIntoView({ behavior: 'smooth', block: 'center' });
            element.classList.add('highlight');
            setTimeout(() => element.classList.remove('highlight'), 1600);
        }

        function renderParagraphText(paragraphText, flaggedSentenceIds, allSentences) {
            // If no flagged sentences, just return escaped text
            if (!flaggedSentenceIds || flaggedSentenceIds.length === 0) {
                return escapeHtml(paragraphText);
            }

            // Build a map of sentence ID to sentence text
            const sentenceMap = {};
            allSentences.forEach(s => {
                if (flaggedSentenceIds.includes(s.id)) {
                    sentenceMap[s.id] = s.bookText;
                }
            });

            // Try to find and hyperlink each flagged sentence in the paragraph
            let result = escapeHtml(paragraphText);

            flaggedSentenceIds.forEach(sentenceId => {
                const sentenceText = sentenceMap[sentenceId];
                if (!sentenceText) return;

                // Escape the sentence text for regex
                const escapedSentenceText = escapeHtml(sentenceText);

                // Find this sentence in the paragraph and wrap it in a link
                // Use word boundary matching to avoid partial matches
                const regex = new RegExp(escapedSentenceText.replace(/[.*+?^${}()|[\\]\\\\]/g, '\\\\\\\\$&'), 'g');
                result = result.replace(regex, `<a href="#sentence-${sentenceId}" onclick="focusSentence(${sentenceId}); return false;" style="color: #4ec9b0; text-decoration: none;">${escapedSentenceText}</a>`);
            });

            return result;
        }

        async function exportAudio(chapterName, startTime, endTime, sentenceId, excerpt) {
            try {
                const response = await fetch(`/api/export/${encodeURIComponent(chapterName)}`, {
                    method: 'POST',
                    headers: {
                        'Content-Type': 'application/json',
                    },
                    body: JSON.stringify({
                        start: startTime,
                        end: endTime,
                        sentenceId: sentenceId,
                        excerpt: excerpt
                    })
                });

                const result = await response.json();

                if (result.success) {
                    showToast(`Audio exported: ${result.filename}<br>Error #${result.errorNumber}`, 'success');
                } else {
                    showToast(`Export failed: ${result.error}`, 'error');
                }
            } catch (err) {
                console.error('Failed to export audio:', err);
                showToast('Failed to export audio. Check console for details.', 'error');
            }
        }

        function openCrxModal(chapterName, startTime, endTime, sentenceId, excerpt) {
            const decodedExcerpt = decodeHtml(excerpt || '');

            pendingCrxData = {
                chapterName: chapterName,
                start: startTime,
                end: endTime,
                sentenceId: sentenceId,
                excerpt: decodedExcerpt
            };

            const commentsField = document.getElementById('comments');
            const sentence = currentReport && Array.isArray(currentReport.sentences)
                ? currentReport.sentences.find(s => s.id === sentenceId)
                : null;

            if (decodedExcerpt) {
                commentsField.value = decodedExcerpt;
            } else if (sentence) {
                const diff = highlightDifferences(sentence.scriptText || '', sentence.bookText || '', sentence.wordOps);
                commentsField.value = `Read as: ${diff.scriptHighlighted}
Should be: ${diff.bookHighlighted}`;
            } else {
                commentsField.value = '';
            }

            document.getElementById('crxModal').style.display = 'block';
        }

        function buildDiffFromWordOps(wordOps, options) {
            if (!Array.isArray(wordOps) || wordOps.length === 0) {
                return null;
            }

            const scriptParts = [];
            const bookParts = [];

            const pushScript = (word, highlighted) => {
                if (!word) return;
                scriptParts.push(highlighted ? options.highlight(word) : options.normal(word));
            };

            const pushBook = (word, highlighted) => {
                if (!word) return;
                bookParts.push(highlighted ? options.highlight(word) : options.normal(word));
            };

            wordOps.forEach(op => {
                const opType = (op.op || '').toLowerCase();
                const scriptWord = (op.asrWord || '').trim();
                const bookWord = (op.bookWord || '').trim();

                switch (opType) {
                    case 'match':
                        pushScript(scriptWord, false);
                        pushBook(bookWord, false);
                        break;
                    case 'sub':
                        pushScript(scriptWord, true);
                        pushBook(bookWord, true);
                        break;
                    case 'del':
                        pushBook(bookWord, true);
                        break;
                    case 'ins':
                        pushScript(scriptWord, true);
                        break;
                    default:
                        pushScript(scriptWord, false);
                        pushBook(bookWord, false);
                        break;
                }
            });

            return {
                script: scriptParts.join(' ').replace(/\\s+/g, ' ').trim(),
                book: bookParts.join(' ').replace(/\\s+/g, ' ').trim()
            };
        }

        function highlightDifferences(script, book, wordOps = null) {
            const diffFromOps = buildDiffFromWordOps(wordOps, {
                normal: word => word,
                highlight: word => `[${word}]`
            });

            if (diffFromOps) {
                return {
                    scriptHighlighted: diffFromOps.script,
                    bookHighlighted: diffFromOps.book
                };
            }

            const scriptWords = script.split(/\\s+/);
            const bookWords = book.split(/\\s+/);

            let scriptResult = [];
            let bookResult = [];
            let scriptMismatchBuffer = [];
            let bookMismatchBuffer = [];

            function normalizeWord(word) {
                return word.replace(/[.,!?;:'"()\\-â€”â€“]/g, '').toLowerCase();
            }

            function flushBuffers() {
                if (scriptMismatchBuffer.length > 0) {
                    scriptResult.push(`[${scriptMismatchBuffer.join(' ')}]`);
                    scriptMismatchBuffer = [];
                }
                if (bookMismatchBuffer.length > 0) {
                    bookResult.push(`[${bookMismatchBuffer.join(' ')}]`);
                    bookMismatchBuffer = [];
                }
            }

            let i = 0, j = 0;

            while (i < scriptWords.length || j < bookWords.length) {
                const scriptWord = scriptWords[i] || '';
                const bookWord = bookWords[j] || '';

                const scriptNorm = normalizeWord(scriptWord);
                const bookNorm = normalizeWord(bookWord);

                if (scriptNorm === bookNorm) {
                    flushBuffers();
                    if (scriptWord) scriptResult.push(scriptWord);
                    if (bookWord) bookResult.push(bookWord);
                    i++;
                    j++;
                } else {
                    let matched = false;
                    for (let n = 1; n <= 3 && i + n <= scriptWords.length; n++) {
                        const scriptPhrase = scriptWords.slice(i, i + n).join('');
                        const scriptPhraseNorm = normalizeWord(scriptPhrase);
                        if (scriptPhraseNorm === bookNorm) {
                            flushBuffers();
                            scriptResult.push(scriptWords.slice(i, i + n).join(' '));
                            bookResult.push(bookWord);
                            i += n;
                            j++;
                            matched = true;
                            break;
                        }
                    }

                    if (!matched) {
                        for (let n = 1; n <= 3 && j + n <= bookWords.length; n++) {
                            const bookPhrase = bookWords.slice(j, j + n).join('');
                            const bookPhraseNorm = normalizeWord(bookPhrase);
                            if (bookPhraseNorm === scriptNorm) {
                                flushBuffers();
                                scriptResult.push(scriptWord);
                                bookResult.push(bookWords.slice(j, j + n).join(' '));
                                i++;
                                j += n;
                                matched = true;
                                break;
                            }
                        }
                    }

                    if (!matched) {
                        if (scriptWord) scriptMismatchBuffer.push(scriptWord);
                        if (bookWord) bookMismatchBuffer.push(bookWord);
                        if (i < scriptWords.length) i++;
                        if (j < bookWords.length) j++;
                    }
                }
            }

            flushBuffers();

            return {
                scriptHighlighted: scriptResult.join(' '),
                bookHighlighted: bookResult.join(' ')
            };
        }


        function highlightDifferencesVisual(script, book, wordOps = null) {
            const diffFromOps = buildDiffFromWordOps(wordOps, {
                normal: word => escapeHtml(word),
                highlight: word => `<mark class="diff-highlight">${escapeHtml(word)}</mark>`
            });

            if (diffFromOps) {
                return {
                    scriptHighlighted: diffFromOps.script,
                    bookHighlighted: diffFromOps.book
                };
            }

            // Word-by-word comparison with visual highlighting using <mark> tags
            const scriptWords = script.split(/\\s+/);
            const bookWords = book.split(/\\s+/);

            let scriptResult = [];
            let bookResult = [];
            let scriptMismatchBuffer = [];
            let bookMismatchBuffer = [];

            function normalizeWord(word) {
                // Remove all punctuation/hyphens/dashes and convert to lowercase for comparison
                // Includes em-dash (â€”), en-dash (â€“), and regular hyphen (-)
                return word.replace(/[.,!?;:'"()\\-â€”â€“]/g, '').toLowerCase();
            }

            function flushBuffers() {
                if (scriptMismatchBuffer.length > 0) {
                    scriptResult.push(`<mark class="diff-highlight">${escapeHtml(scriptMismatchBuffer.join(' '))}</mark>`);
                    scriptMismatchBuffer = [];
                }
                if (bookMismatchBuffer.length > 0) {
                    bookResult.push(`<mark class="diff-highlight">${escapeHtml(bookMismatchBuffer.join(' '))}</mark>`);
                    bookMismatchBuffer = [];
                }
            }

            let i = 0, j = 0;

            while (i < scriptWords.length || j < bookWords.length) {
                const scriptWord = scriptWords[i] || '';
                const bookWord = bookWords[j] || '';

                const scriptNorm = normalizeWord(scriptWord);
                const bookNorm = normalizeWord(bookWord);

                if (scriptNorm === bookNorm) {
                    flushBuffers();
                    if (scriptWord) scriptResult.push(escapeHtml(scriptWord));
                    if (bookWord) bookResult.push(escapeHtml(bookWord));
                    i++;
                    j++;
                } else {
                    let matched = false;
                    for (let n = 1; n <= 3 && i + n <= scriptWords.length; n++) {
                        const scriptPhrase = scriptWords.slice(i, i + n).join('');
                        const scriptPhraseNorm = normalizeWord(scriptPhrase);
                        if (scriptPhraseNorm === bookNorm) {
                            flushBuffers();
                            scriptResult.push(escapeHtml(scriptWords.slice(i, i + n).join(' ')));
                            bookResult.push(escapeHtml(bookWord));
                            i += n;
                            j++;
                            matched = true;
                            break;
                        }
                    }

                    if (!matched) {
                        for (let n = 1; n <= 3 && j + n <= bookWords.length; n++) {
                            const bookPhrase = bookWords.slice(j, j + n).join('');
                            const bookPhraseNorm = normalizeWord(bookPhrase);
                            if (bookPhraseNorm === scriptNorm) {
                                flushBuffers();
                                scriptResult.push(escapeHtml(scriptWord));
                                bookResult.push(escapeHtml(bookWords.slice(j, j + n).join(' ')));
                                i++;
                                j += n;
                                matched = true;
                                break;
                            }
                        }
                    }

                    if (!matched) {
                        if (scriptWord) scriptMismatchBuffer.push(scriptWord);
                        if (bookWord) bookMismatchBuffer.push(bookWord);
                        if (i < scriptWords.length) i++;
                        if (j < bookWords.length) j++;
                    }
                }
            }

            flushBuffers();

            return {
                scriptHighlighted: scriptResult.join(' '),
                bookHighlighted: bookResult.join(' ')
            };
        }

        function renderUnifiedDiff(diff) {
            // Render the transcript (what was read) with inline diff highlighting
            // Deletions = what was in the manuscript but NOT read (strikethrough red)
            // Insertions = what was actually read but NOT in manuscript (green highlight)
            if (!diff || !diff.ops || !Array.isArray(diff.ops)) {
                return '<div class="diff-empty">No diff data available</div>';
            }

            const parts = [];

            diff.ops.forEach(op => {
                const tokens = op.tokens || [];

                if (tokens.length === 0) return;

                const text = tokens.join(' ');

                if (op.op === 'equal') {
                    // Equal tokens - what was read correctly
                    parts.push(escapeHtml(text));
                } else if (op.op === 'delete') {
                    // Deletion - what should have been read but wasn't (strikethrough)
                    parts.push(`<span class="diff-deleted">${escapeHtml(text)}</span>`);
                } else if (op.op === 'insert') {
                    // Insertion - what was actually read (not in manuscript)
                    parts.push(`<span class="diff-inserted">${escapeHtml(text)}</span>`);
                }
            });

            if (parts.length === 0) {
                return '<div class="diff-empty">No transcript data</div>';
            }

            return '<div class="diff-inline">' + parts.join(' ') + '</div>';
        }

        function closeCrxModal() {
            document.getElementById('crxModal').style.display = 'none';
            pendingCrxData = null;
        }

        async function submitCrx() {
            if (!pendingCrxData) return;

            const errorType = document.getElementById('errorType').value;
            const comments = document.getElementById('comments').value;

            try {
                const response = await fetch(`/api/crx/${encodeURIComponent(pendingCrxData.chapterName)}`, {
                    method: 'POST',
                    headers: {
                        'Content-Type': 'application/json',
                    },
                    body: JSON.stringify({
                        start: pendingCrxData.start,
                        end: pendingCrxData.end,
                        sentenceId: pendingCrxData.sentenceId,
                        errorType: errorType,
                        comments: comments,
                        excerpt: pendingCrxData.excerpt
                    })
                });

                const result = await response.json();

                if (result.success) {
                    showToast(`Added to CRX!<br>Error #${result.errorNumber} â€¢ ${result.timecode}`, 'success');
                    closeCrxModal();
                } else {
                    showToast(`Failed to add to CRX: ${result.error}`, 'error');
                }
            } catch (err) {
                console.error('Failed to add to CRX:', err);
                showToast('Failed to add to CRX. Check console for details.', 'error');
            }
        }

        // Toast notification system
        function showToast(message, type = 'success', duration = 5000) {
            const container = document.getElementById('toast-container');
            const toast = document.createElement('div');
            toast.className = `toast ${type}`;

            const title = type === 'success' ? 'Success' : 'Error';

            toast.innerHTML = `
                <div class="toast-header">
                    <div class="toast-title">${title}</div>
                    <button class="toast-close" onclick="this.parentElement.parentElement.remove()">Ã—</button>
                </div>
                <div class="toast-message">${message}</div>
            `;

            container.appendChild(toast);

            // Auto-remove after duration
            setTimeout(() => {
                toast.style.animation = 'slideIn 0.3s ease-out reverse';
                setTimeout(() => toast.remove(), 300);
            }, duration);
        }

        // Close modal when clicking outside of it
        window.onclick = function(event) {
            const modal = document.getElementById('crxModal');
            if (event.target === modal) {
                closeCrxModal();
            }
        }

        // Load chapters and overview on page load
        loadChapters().then(() => {
            loadOverview();
        });
    </script>
</body>
</html>"""

        self.send_response(200)
        self.send_header('Content-type', 'text/html')
        self.end_headers()
        self.wfile.write(html.encode())

    def extract_hydrate_metrics(self, hydrate_path):
        """Extract summary metrics from hydrate.json"""
        try:
            with open(hydrate_path, 'r', encoding='utf-8') as f:
                hydrate_data = json.load(f)

            sentences = hydrate_data.get('sentences', [])
            paragraphs = hydrate_data.get('paragraphs', [])

            # Count flagged sentences
            flagged_sentences = [s for s in sentences if s.get('status', 'ok') != 'ok']

            # Calculate average WER
            total_wer = sum(s.get('metrics', {}).get('wer', 0) for s in sentences if 'metrics' in s)
            avg_wer = (total_wer / len(sentences) * 100) if sentences else 0

            return {
                'sentenceCount': len(sentences),
                'sentenceFlagged': len(flagged_sentences),
                'sentenceAvgWer': f"{avg_wer:.2f}%",
                'paragraphCount': len(paragraphs),
                'paragraphFlagged': 0,
                'paragraphAvgWer': '0.00%'
            }
        except Exception as e:
            print(f"Error extracting metrics from {hydrate_path}: {e}")
            return {
                'sentenceCount': 0,
                'sentenceFlagged': 0,
                'sentenceAvgWer': 'N/A',
                'paragraphCount': 0,
                'paragraphFlagged': 0,
                'paragraphAvgWer': 'N/A'
            }

    def extract_report_metrics(self, report_path):
        """Extract summary metrics from a validation report"""
        try:
            with open(report_path, 'r', encoding='utf-8-sig') as f:
                content = f.read()

            metrics = {
                'sentenceCount': 0,
                'sentenceFlagged': 0,
                'sentenceAvgWer': '0.00%',
                'paragraphCount': 0,
                'paragraphFlagged': 0,
                'paragraphAvgWer': '0.00%'
            }

            # Parse sentences line: "Sentences : 277 (Avg WER 2.48%, Max WER 100.00%, Flagged 18)"
            sent_match = re.search(r'Sentences\s*:\s*(\d+)\s*\(Avg WER ([\d.]+)%.*Flagged (\d+)\)', content)
            if sent_match:
                metrics['sentenceCount'] = int(sent_match.group(1))
                metrics['sentenceAvgWer'] = sent_match.group(2) + '%'
                metrics['sentenceFlagged'] = int(sent_match.group(3))

            # Parse paragraphs line: "Paragraphs: 72 (Avg WER 3.57%, Avg Coverage 99.89%)"
            para_match = re.search(r'Paragraphs\s*:\s*(\d+)\s*\(Avg WER ([\d.]+)%', content)
            if para_match:
                metrics['paragraphCount'] = int(para_match.group(1))
                metrics['paragraphAvgWer'] = para_match.group(2) + '%'

            # Count flagged paragraphs
            para_section = content.split('All paragraphs by WER:')
            if len(para_section) > 1:
                flagged_paras = re.findall(r'#\d+ \| WER [\d.]+% \| Coverage [\d.]+% \| Status (attention|unreliable)', para_section[1])
                metrics['paragraphFlagged'] = len(flagged_paras)

            return metrics
        except Exception as e:
            print(f"Error extracting metrics from {report_path}: {e}")
            return {
                'sentenceCount': 0,
                'sentenceFlagged': 0,
                'sentenceAvgWer': 'N/A',
                'paragraphCount': 0,
                'paragraphFlagged': 0,
                'paragraphAvgWer': 'N/A'
            }

    def serve_chapters_list(self):
        """Find all hydrate files and return chapter list with metrics"""
        chapters = []

        for item in BASE_DIR.iterdir():
            if item.is_dir():
                # Look for hydrate.json in this directory
                hydrate_pattern = item.name + ".align.hydrate.json"
                hydrate_path = item / hydrate_pattern

                if hydrate_path.exists():
                    # Extract basic metrics from hydrate
                    metrics = self.extract_hydrate_metrics(hydrate_path)
                    chapters.append({
                        'name': item.name,
                        'path': str(item.relative_to(BASE_DIR)),
                        'metrics': metrics
                    })

        def natural_sort_key(chapter):
            name = chapter['name']
            number_groups = [int(match.group()) for match in re.finditer(r"\d+", name)]
            if number_groups:
                # Prefer the first number block (chapters typically lead with it)
                primary = number_groups[0]
                # Secondary key ensures ties keep lexicographic order
                return (0, primary, name.lower())
            return (1, name.lower())

        chapters.sort(key=natural_sort_key)

        self.send_response(200)
        self.send_header('Content-type', 'application/json')
        self.end_headers()
        self.wfile.write(json.dumps(chapters).encode())

    def serve_overview(self):
        """Serve book-wide overview with aggregated metrics"""
        chapters = []

        for item in BASE_DIR.iterdir():
            if item.is_dir():
                hydrate_pattern = item.name + ".align.hydrate.json"
                hydrate_path = item / hydrate_pattern

                if hydrate_path.exists():
                    metrics = self.extract_hydrate_metrics(hydrate_path)
                    chapters.append({
                        'name': item.name,
                        'path': str(item.relative_to(BASE_DIR)),
                        'metrics': metrics
                    })

        def natural_sort_key(chapter):
            name = chapter['name']
            number_groups = [int(match.group()) for match in re.finditer(r"\d+", name)]
            if number_groups:
                primary = number_groups[0]
                return (0, primary, name.lower())
            return (1, name.lower())

        chapters.sort(key=natural_sort_key)

        # Calculate book-wide totals
        total_sentences = sum(c['metrics']['sentenceCount'] for c in chapters)
        total_flagged_sentences = sum(c['metrics']['sentenceFlagged'] for c in chapters)
        total_paragraphs = sum(c['metrics']['paragraphCount'] for c in chapters)
        total_flagged_paragraphs = sum(c['metrics']['paragraphFlagged'] for c in chapters)

        # Calculate weighted average WER
        if total_sentences > 0:
            weighted_sent_wer_sum = sum(
                c['metrics']['sentenceCount'] * float(c['metrics']['sentenceAvgWer'].rstrip('%'))
                for c in chapters if c['metrics']['sentenceAvgWer'] != 'N/A'
            )
            avg_sentence_wer = weighted_sent_wer_sum / total_sentences
        else:
            avg_sentence_wer = 0

        if total_paragraphs > 0:
            weighted_para_wer_sum = sum(
                c['metrics']['paragraphCount'] * float(c['metrics']['paragraphAvgWer'].rstrip('%'))
                for c in chapters if c['metrics']['paragraphAvgWer'] != 'N/A'
            )
            avg_paragraph_wer = weighted_para_wer_sum / total_paragraphs
        else:
            avg_paragraph_wer = 0

        overview = {
            'bookName': BASE_DIR.name,
            'chapterCount': len(chapters),
            'totalSentences': total_sentences,
            'totalFlaggedSentences': total_flagged_sentences,
            'avgSentenceWer': f"{avg_sentence_wer:.2f}%",
            'totalParagraphs': total_paragraphs,
            'totalFlaggedParagraphs': total_flagged_paragraphs,
            'avgParagraphWer': f"{avg_paragraph_wer:.2f}%",
            'chapters': chapters
        }

        self.send_response(200)
        self.send_header('Content-type', 'application/json')
        self.end_headers()
        self.wfile.write(json.dumps(overview).encode())

    def serve_report(self, chapter_name):
        """Parse and serve a validation report from hydrate.json"""
        from urllib.parse import unquote

        # Decode URL-encoded chapter name
        chapter_name = unquote(chapter_name)

        chapter_dir = BASE_DIR / chapter_name
        hydrate_file = chapter_dir / f"{chapter_name}.align.hydrate.json"

        print(f"Looking for hydrate: {hydrate_file}")
        print(f"Hydrate exists: {hydrate_file.exists()}")

        if not hydrate_file.exists():
            error_msg = json.dumps({
                'error': 'Hydrate file not found',
                'chapter': chapter_name,
                'path': str(hydrate_file)
            })
            self.send_response(404)
            self.send_header('Content-type', 'application/json')
            self.end_headers()
            self.wfile.write(error_msg.encode())
            return

        try:
            # Load hydrate data
            with open(hydrate_file, 'r', encoding='utf-8') as f:
                hydrate_data = json.load(f)

            report_data = self.parse_hydrate_to_report(hydrate_data, chapter_name)
            self.send_response(200)
            self.send_header('Content-type', 'application/json')
            self.end_headers()
            self.wfile.write(json.dumps(report_data).encode())
        except Exception as e:
            import traceback
            print(f"Error parsing hydrate: {e}")
            print(traceback.format_exc())

            error_msg = json.dumps({
                'error': str(e),
                'traceback': traceback.format_exc()
            })
            self.send_response(500)
            self.send_header('Content-type', 'application/json')
            self.end_headers()
            self.wfile.write(error_msg.encode())

    def parse_hydrate_to_report(self, hydrate_data, chapter_name):
        """Convert hydrate.json data to report format expected by the UI"""
        sentences = hydrate_data.get('sentences', [])
        paragraphs = hydrate_data.get('paragraphs', [])

        # Calculate statistics
        flagged_sentences = [s for s in sentences if s.get('status', 'ok') != 'ok']
        total_wer = sum(s.get('metrics', {}).get('wer', 0) for s in sentences if 'metrics' in s)
        avg_wer = (total_wer / len(sentences) * 100) if sentences else 0
        max_wer = max((s.get('metrics', {}).get('wer', 0) for s in sentences if 'metrics' in s), default=0) * 100

        # Convert sentences to UI format
        ui_sentences = []
        for sent in sentences:
            metrics = sent.get('metrics', {})
            timing = sent.get('timing', {})
            book_range = sent.get('bookRange', {})
            script_range = sent.get('scriptRange', {})

            wer = metrics.get('wer', 0) * 100
            cer = metrics.get('cer', 0) * 100

            ui_sent = {
                'id': sent.get('id'),
                'wer': f"{wer:.1f}%",
                'cer': f"{cer:.1f}%",
                'status': sent.get('status', 'ok'),
                'bookRange': f"{book_range.get('start', 0)}-{book_range.get('end', 0)}",
                'scriptRange': f"{script_range.get('start', 0)}-{script_range.get('end', 0)}",
                'timing': f"{timing.get('startSec', 0):.3f}s â†’ {timing.get('endSec', 0):.3f}s (Î” {timing.get('duration', 0):.3f}s)",
                'bookText': sent.get('bookText', ''),
                'scriptText': sent.get('scriptText', ''),
                'excerpt': sent.get('bookText', '')[:100],
                'diff': sent.get('diff', {}),  # Include the full diff structure
                'startTime': timing.get('startSec', 0),
                'endTime': timing.get('endSec', 0),
                'bookRangeStart': book_range.get('start'),
                'bookRangeEnd': book_range.get('end')
            }
            ui_sentences.append(ui_sent)

        # Sort by WER descending
        ui_sentences.sort(key=lambda s: float(s['wer'].rstrip('%')), reverse=True)

        # Build report data structure
        report_data = {
            'chapterName': chapter_name,
            'audioPath': hydrate_data.get('audioPath', ''),
            'scriptPath': '',
            'bookIndex': '',
            'created': datetime.now().isoformat(),
            'stats': {
                'sentenceCount': str(len(sentences)),
                'avgWer': f"{avg_wer:.2f}%",
                'maxWer': f"{max_wer:.2f}%",
                'flaggedCount': str(len(flagged_sentences)),
                'paragraphCount': str(len(paragraphs)),
                'paragraphAvgWer': '0.00%',
                'avgCoverage': '100.00%'
            },
            'sentences': ui_sentences,
            'paragraphs': []
        }

        return report_data

    def parse_report(self, report_path, chapter_name, hydrate_data=None):
        """Parse validation report text file into structured data"""
        with open(report_path, 'r', encoding='utf-8-sig') as f:
            content = f.read()

        lines = content.split('\n')

        # Parse header
        report_data = {
            'chapterName': chapter_name,
            'audioPath': '',
            'scriptPath': '',
            'bookIndex': '',
            'created': '',
            'stats': {},
            'sentences': [],
            'paragraphs': []
        }

        def parse_index_range(range_text):
            if not range_text:
                return None, None

            match = re.search(r'(\d+)\s*-\s*(\d+)', range_text)
            if match:
                return int(match.group(1)), int(match.group(2))

            match = re.search(r'(\d+)', range_text)
            if match:
                value = int(match.group(1))
                return value, value

            return None, None

        # Extract metadata from header
        for line in lines[:10]:
            if line.startswith('Audio'):
                report_data['audioPath'] = line.split(':', 1)[1].strip()
            elif line.startswith('Script'):
                report_data['scriptPath'] = line.split(':', 1)[1].strip()
            elif line.startswith('Book Index'):
                report_data['bookIndex'] = line.split(':', 1)[1].strip()
            elif line.startswith('Created'):
                report_data['created'] = line.split(':', 1)[1].strip()
            elif line.startswith('Sentences'):
                # Parse: "Sentences : 277 (Avg WER 2.30%, Max WER 100.00%, Flagged 19)"
                match = re.search(r'(\d+)\s*\(Avg WER ([\d.]+)%, Max WER ([\d.]+)%, Flagged (\d+)\)', line)
                if match:
                    report_data['stats']['sentenceCount'] = match.group(1)
                    report_data['stats']['avgWer'] = match.group(2) + '%'
                    report_data['stats']['maxWer'] = match.group(3) + '%'
                    report_data['stats']['flaggedCount'] = match.group(4)
            elif line.startswith('Paragraphs'):
                # Parse: "Paragraphs: 72 (Avg WER 3.53%, Avg Coverage 98.93%)"
                match = re.search(r'(\d+)\s*\(Avg WER ([\d.]+)%, Avg Coverage ([\d.]+)%\)', line)
                if match:
                    report_data['stats']['paragraphCount'] = match.group(1)
                    report_data['stats']['paragraphAvgWer'] = match.group(2) + '%'
                    report_data['stats']['avgCoverage'] = match.group(3) + '%'

        # Parse sentences section
        current_section = None
        current_item = None

        for i, line in enumerate(lines):
            line = line.strip()

            if line == 'All sentences by WER:':
                current_section = 'sentences'
                continue
            elif line == 'All paragraphs by WER:':
                current_section = 'paragraphs'
                continue

            if current_section == 'sentences':
                # Match sentence header: "  #43 | WER 100.0% | CER 100.0% | Status unreliable"
                match = re.match(r'#(\d+)\s*\|\s*WER\s*([\d.]+)%\s*\|\s*CER\s*([\d.]+)%\s*\|\s*Status\s*(\w+)', line)
                if match:
                    if current_item:
                        report_data['sentences'].append(current_item)

                    current_item = {
                        'id': int(match.group(1)),
                        'wer': match.group(2) + '%',
                        'cer': match.group(3) + '%',
                        'status': match.group(4),
                        'bookRange': '',
                        'scriptRange': '',
                        'timing': '',
                        'bookText': '',
                        'scriptText': '',
                        'excerpt': '',
                        'startTime': None,
                        'endTime': None,
                        'bookRangeStart': None,
                        'bookRangeEnd': None
                    }
                elif current_item:
                    if line.startswith('Book range:'):
                        range_text = line.split(':', 1)[1].strip()
                        current_item['bookRange'] = range_text
                        start, end = parse_index_range(range_text)
                        current_item['bookRangeStart'] = start
                        current_item['bookRangeEnd'] = end
                    elif line.startswith('Script range:'):
                        current_item['scriptRange'] = line.split(':', 1)[1].strip()
                    elif line.startswith('Timing:'):
                        timing_str = line.split(':', 1)[1].strip()
                        current_item['timing'] = timing_str
                        # Parse timing: "870.530s â†’ 871.050s (Î” 0.520s)"
                        timing_match = re.search(r'([\d.]+)s\s*â†’\s*([\d.]+)s', timing_str)
                        if timing_match:
                            current_item['startTime'] = float(timing_match.group(1))
                            current_item['endTime'] = float(timing_match.group(2))
                    elif line.startswith('Book   :'):
                        current_item['bookText'] = line.split(':', 1)[1].strip()
                    elif line.startswith('Script :'):
                        current_item['scriptText'] = line.split(':', 1)[1].strip()
                    elif line.startswith('Excerpt:'):
                        current_item['excerpt'] = line.split(':', 1)[1].strip()

            elif current_section == 'paragraphs':
                # Match paragraph header: "  #44 | WER 100.0% | Coverage 100.0% | Status unreliable"
                match = re.match(r'#(\d+)\s*\|\s*WER\s*([\d.]+)%\s*\|\s*Coverage\s*([\d.]+)%\s*\|\s*Status\s*(\w+)', line)
                if match:
                    if current_item and 'coverage' in current_item:
                        report_data['paragraphs'].append(current_item)

                    current_item = {
                        'id': int(match.group(1)),
                        'wer': match.group(2) + '%',
                        'coverage': match.group(3) + '%',
                        'status': match.group(4),
                        'bookRange': '',
                        'bookText': '',
                        'bookRangeStart': None,
                        'bookRangeEnd': None,
                        'sentenceIds': [],
                        'startTime': None,
                        'endTime': None,
                        'timing': ''
                    }
                elif current_item and 'coverage' in current_item:
                    if line.startswith('Book range:'):
                        range_text = line.split(':', 1)[1].strip()
                        current_item['bookRange'] = range_text
                        start, end = parse_index_range(range_text)
                        current_item['bookRangeStart'] = start
                        current_item['bookRangeEnd'] = end
                    elif line.startswith('Book   :'):
                        current_item['bookText'] = line.split(':', 1)[1].strip()

        # Add last item
        if current_item:
            if current_section == 'sentences':
                report_data['sentences'].append(current_item)
            elif current_section == 'paragraphs' and 'coverage' in current_item:
                report_data['paragraphs'].append(current_item)

        # Map paragraphs to sentences for linking and audio timing
        # Use hydrate data if available to get complete sentence list with timing
        if hydrate_data and 'paragraphs' in hydrate_data and 'sentences' in hydrate_data:
            # Create lookup for hydrate sentences by ID
            hydrate_sentences = {s['id']: s for s in hydrate_data['sentences']}
            hydrate_paragraphs = {p['id']: p for p in hydrate_data['paragraphs']}

            # Build a map from sentence ID to paragraph ID
            sentence_to_paragraph = {}
            for para in hydrate_data['paragraphs']:
                for sent_id in para.get('sentenceIds', []):
                    sentence_to_paragraph[sent_id] = para['id']

            # Build alignment word operations grouped by sentence, preserving order
            word_ops_by_sentence = {}
            words = hydrate_data.get('words') or []
            if words:
                # Map each book word index to its sentence ID for quick lookup
                book_to_sentence = {}
                for sent in hydrate_data['sentences']:
                    book_range = sent.get('bookRange') or {}
                    start = book_range.get('start')
                    end = book_range.get('end')
                    if start is None or end is None:
                        continue
                    if end < start:
                        start, end = end, start

                    for idx in range(start, end + 1):
                        book_to_sentence[idx] = sent['id']

                    word_ops_by_sentence.setdefault(sent['id'], [])

                last_sentence_id = None
                for word in words:
                    sentence_id = None
                    book_idx = word.get('bookIdx')
                    if book_idx is not None:
                        sentence_id = book_to_sentence.get(book_idx)
                    if sentence_id is None:
                        sentence_id = last_sentence_id
                    if sentence_id is None:
                        continue

                    entry = {
                        'op': word.get('op'),
                        'reason': word.get('reason'),
                        'bookWord': (word.get('bookWord') or '').strip(),
                        'asrWord': (word.get('asrWord') or '').strip()
                    }
                    word_ops_by_sentence.setdefault(sentence_id, []).append(entry)
                    last_sentence_id = sentence_id

            # Add paragraph ID to each sentence
            for sentence in report_data['sentences']:
                sentence['paragraphId'] = sentence_to_paragraph.get(sentence['id'])
                if word_ops_by_sentence:
                    ops = word_ops_by_sentence.get(sentence['id'])
                    if ops:
                        sentence['wordOps'] = ops

            for paragraph in report_data['paragraphs']:
                para_id = paragraph.get('id')
                hydrate_para = hydrate_paragraphs.get(para_id)

                if hydrate_para and 'sentenceIds' in hydrate_para:
                    # Get ALL sentence IDs from hydrate (not just flagged ones)
                    all_sentence_ids = hydrate_para['sentenceIds']
                    paragraph['sentenceIds'] = all_sentence_ids

                    # Find which of these sentences are flagged (in report_data['sentences'])
                    flagged_ids = {s['id'] for s in report_data['sentences']}
                    paragraph['flaggedSentenceIds'] = [sid for sid in all_sentence_ids if sid in flagged_ids]

                    # Calculate timing from ALL sentences in paragraph
                    # Also include last sentence of previous paragraph and first sentence of next paragraph
                    context_sentence_ids = all_sentence_ids.copy()

                    # Add last sentence from previous paragraph (by ID, not by sorted position)
                    prev_para_id = para_id - 1
                    if prev_para_id >= 0:
                        prev_hydrate = hydrate_paragraphs.get(prev_para_id)
                        if prev_hydrate and 'sentenceIds' in prev_hydrate and prev_hydrate['sentenceIds']:
                            last_prev_sent_id = prev_hydrate['sentenceIds'][-1]
                            if last_prev_sent_id not in context_sentence_ids:
                                context_sentence_ids.insert(0, last_prev_sent_id)

                    # Add first sentence from next paragraph (by ID, not by sorted position)
                    next_para_id = para_id + 1
                    next_hydrate = hydrate_paragraphs.get(next_para_id)
                    if next_hydrate and 'sentenceIds' in next_hydrate and next_hydrate['sentenceIds']:
                        first_next_sent_id = next_hydrate['sentenceIds'][0]
                        if first_next_sent_id not in context_sentence_ids:
                            context_sentence_ids.append(first_next_sent_id)

                    start_times = []
                    end_times = []
                    for sent_id in context_sentence_ids:
                        hydrate_sent = hydrate_sentences.get(sent_id)
                        if hydrate_sent and 'timing' in hydrate_sent:
                            timing = hydrate_sent['timing']
                            if 'startSec' in timing and 'endSec' in timing:
                                start_times.append(timing['startSec'])
                                end_times.append(timing['endSec'])

                    if start_times and end_times:
                        paragraph_start = min(start_times)
                        paragraph_end = max(end_times)
                        paragraph['startTime'] = paragraph_start
                        paragraph['endTime'] = paragraph_end
                        duration = paragraph_end - paragraph_start
                        paragraph['timing'] = f"{paragraph_start:.3f}s â†’ {paragraph_end:.3f}s (Î” {duration:.3f}s)"
                    else:
                        paragraph['startTime'] = None
                        paragraph['endTime'] = None
                        paragraph['timing'] = ''
                else:
                    # Fallback if hydrate data not available for this paragraph
                    paragraph['sentenceIds'] = []
                    paragraph['flaggedSentenceIds'] = []
                    paragraph['startTime'] = None
                    paragraph['endTime'] = None
                    paragraph['timing'] = ''
        else:
            # Original fallback logic using only flagged sentences
            for paragraph in report_data['paragraphs']:
                start = paragraph.get('bookRangeStart')
                end = paragraph.get('bookRangeEnd')
                if start is None or end is None:
                    paragraph['sentenceIds'] = []
                    paragraph['flaggedSentenceIds'] = []
                    paragraph['startTime'] = None
                    paragraph['endTime'] = None
                    paragraph['timing'] = ''
                    continue

                sentence_ids = []
                start_times = []
                end_times = []

                for sentence in report_data['sentences']:
                    sentence_start = sentence.get('bookRangeStart')
                    sentence_end = sentence.get('bookRangeEnd')
                    if sentence_start is None or sentence_end is None:
                        continue

                    if sentence_end < start or sentence_start > end:
                        continue

                    sentence_ids.append(sentence['id'])
                    if sentence.get('startTime') is not None:
                        start_times.append(sentence['startTime'])
                    if sentence.get('endTime') is not None:
                        end_times.append(sentence['endTime'])

                sentence_ids.sort()
                paragraph['sentenceIds'] = sentence_ids
                paragraph['flaggedSentenceIds'] = sentence_ids  # All are flagged in fallback mode

                if start_times and end_times:
                    paragraph_start = min(start_times)
                    paragraph_end = max(end_times)
                    if paragraph_end < paragraph_start:
                        paragraph_end = paragraph_start
                    paragraph['startTime'] = paragraph_start
                    paragraph['endTime'] = paragraph_end
                    duration = paragraph_end - paragraph_start
                    paragraph['timing'] = f"{paragraph_start:.3f}s â†’ {paragraph_end:.3f}s (Î” {duration:.3f}s)"
                else:
                    paragraph['startTime'] = None
                    paragraph['endTime'] = None
                    paragraph['timing'] = ''

        return report_data

    def serve_audio_segment(self, chapter_name, query_string):
        """Serve an audio segment for a specific chapter and time range"""
        from urllib.parse import unquote, parse_qs
        import subprocess
        import tempfile

        # Parse chapter name and query parameters
        chapter_name = unquote(chapter_name)
        query_params = parse_qs(query_string) if query_string else {}

        start_time = float(query_params.get('start', [0])[0])
        end_time = float(query_params.get('end', [0])[0])

        # Find the audio file - try chapter folder first, then book root
        audio_path = BASE_DIR / chapter_name / f"{chapter_name}.treated.wav"
        if not audio_path.exists():
            audio_path = BASE_DIR / chapter_name / f"{chapter_name}.wav"
        if not audio_path.exists():
            audio_path = BASE_DIR / f"{chapter_name}.wav"

        print(f"Looking for audio: {audio_path}")
        print(f"Audio exists: {audio_path.exists()}")
        print(f"Segment: {start_time}s to {end_time}s")

        if not audio_path.exists():
            error_msg = json.dumps({
                'error': 'Audio file not found',
                'path': str(audio_path)
            })
            self.send_response(404)
            self.send_header('Content-type', 'application/json')
            self.end_headers()
            self.wfile.write(error_msg.encode())
            return

        try:
            # Use ffmpeg to extract the audio segment
            duration = end_time - start_time
            with tempfile.NamedTemporaryFile(suffix='.wav', delete=False) as temp_file:
                temp_path = temp_file.name

            # Extract segment using ffmpeg
            cmd = [
                'ffmpeg',
                '-i', str(audio_path),
                '-ss', str(start_time),
                '-t', str(duration),
                '-c', 'copy',
                '-y',
                temp_path
            ]

            result = subprocess.run(cmd, capture_output=True, text=True)

            if result.returncode != 0:
                raise Exception(f"ffmpeg failed: {result.stderr}")

            # Serve the file
            with open(temp_path, 'rb') as f:
                audio_data = f.read()

            self.send_response(200)
            self.send_header('Content-type', 'audio/wav')
            self.send_header('Content-Length', str(len(audio_data)))
            self.end_headers()
            self.wfile.write(audio_data)

            # Clean up temp file
            os.unlink(temp_path)

        except Exception as e:
            import traceback
            print(f"Error serving audio: {e}")
            print(traceback.format_exc())

            error_msg = json.dumps({
                'error': str(e),
                'traceback': traceback.format_exc()
            })
            self.send_response(500)
            self.send_header('Content-type', 'application/json')
            self.end_headers()
            self.wfile.write(error_msg.encode())

    def handle_export_audio(self, chapter_name):
        """Export audio segment to CRX folder"""
        from urllib.parse import unquote

        chapter_name = unquote(chapter_name)
        print(f"Export audio request for chapter: {chapter_name}")

        # Read POST data
        content_length = int(self.headers['Content-Length'])
        post_data = self.rfile.read(content_length)
        params = json.loads(post_data.decode('utf-8'))

        start_time = float(params['start'])
        end_time = float(params['end'])
        sentence_id = params.get('sentenceId', 'unknown')

        # Find the audio file - try chapter folder first, then book root
        audio_path = BASE_DIR / chapter_name / f"{chapter_name}.treated.wav"
        if not audio_path.exists():
            audio_path = BASE_DIR / chapter_name / f"{chapter_name}.wav"
        if not audio_path.exists():
            audio_path = BASE_DIR / f"{chapter_name}.wav"

        print(f"Looking for audio at: {audio_path}")
        print(f"Audio exists: {audio_path.exists()}")

        if not audio_path.exists():
            self.send_json_response({'error': 'Audio file not found'}, 404)
            return

        try:
            # Create CRX directory if it doesn't exist
            crx_dir = BASE_DIR / CRX_DIR_NAME
            crx_dir.mkdir(exist_ok=True)

            # Determine the next error number by checking existing audio files
            # Find the highest numbered .wav file in the CRX directory
            existing_files = list(crx_dir.glob('*.wav'))
            if existing_files:
                # Extract numbers from filenames like "001.wav", "002.wav"
                error_numbers = []
                for f in existing_files:
                    match = re.match(r'(\d+)\.wav', f.name)
                    if match:
                        error_numbers.append(int(match.group(1)))

                if error_numbers:
                    error_num = max(error_numbers) + 1
                else:
                    error_num = 1
            else:
                error_num = 1

            # Generate filename: ErrorNum.wav (e.g., 001.wav)
            export_filename = f"{error_num:03d}.wav"
            export_path = crx_dir / export_filename

            # Extract segment using ffmpeg
            duration = end_time - start_time
            cmd = [
                'ffmpeg',
                '-i', str(audio_path),
                '-ss', str(start_time),
                '-t', str(duration),
                '-c', 'copy',
                '-y',
                str(export_path)
            ]

            result = subprocess.run(cmd, capture_output=True, text=True)

            if result.returncode != 0:
                raise Exception(f"ffmpeg failed: {result.stderr}")

            self.send_json_response({
                'success': True,
                'filename': export_filename,
                'errorNumber': error_num,
                'path': str(export_path.relative_to(BASE_DIR))
            })

        except Exception as e:
            print(f"Error exporting audio: {e}")
            print(traceback.format_exc())
            self.send_json_response({'error': str(e)}, 500)

    def handle_add_to_crx(self, chapter_name):
        """Add entry to CRX Excel file"""
        from urllib.parse import unquote

        chapter_name = unquote(chapter_name)

        # Read POST data
        content_length = int(self.headers['Content-Length'])
        post_data = self.rfile.read(content_length)
        params = json.loads(post_data.decode('utf-8'))

        start_time = float(params['start'])
        end_time = float(params['end'])
        sentence_id = params.get('sentenceId', 'unknown')
        error_type = params.get('errorType', DEFAULT_ERROR_TYPE)
        comments = params.get('comments', '')

        try:
            # Create CRX directory if it doesn't exist
            crx_dir = BASE_DIR / CRX_DIR_NAME
            crx_dir.mkdir(exist_ok=True)

            # CRX file path
            crx_filename = f"{BASE_DIR.name}_CRX.xlsx"
            crx_path = crx_dir / crx_filename

            # Determine error number by checking existing audio files
            existing_files = list(crx_dir.glob('*.wav'))
            if existing_files:
                error_numbers = []
                for f in existing_files:
                    match = re.match(r'(\d+)\.wav', f.name)
                    if match:
                        error_numbers.append(int(match.group(1)))
                error_num = max(error_numbers) if error_numbers else 1
            else:
                error_num = 1

            # If CRX doesn't exist, copy BASE_CRX template
            if not crx_path.exists():
                if not CRX_TEMPLATE_PATH.exists():
                    raise Exception(f"CRX template not found: {CRX_TEMPLATE_PATH}")
                # Use shutil.copy2 to preserve metadata and formatting
                shutil.copy2(CRX_TEMPLATE_PATH, crx_path)

            # Use openpyxl to modify cells
            wb = openpyxl.load_workbook(crx_path)
            ws = wb.active

            # Find the row corresponding to this error number
            # Row number = CRX_DATA_ROW_START + (error_num - 1)
            target_row = CRX_DATA_ROW_START + (error_num - 1)

            # Format timecode as hh:mm:ss
            hours = int(start_time // 3600)
            minutes = int((start_time % 3600) // 60)
            seconds = int(start_time % 60)
            timecode = f"{hours:02d}:{minutes:02d}:{seconds:02d}"

            # Ensure Error # is populated in this row (Column B = 2)
            current_error = ws.cell(row=target_row, column=2).value
            if not current_error or not re.search(r'(\d+)', str(current_error)):
                ws.cell(row=target_row, column=2).value = f"{error_num:03d}"

            # Write data to cells
            # Column C: Recording Day (leave empty)
            # Column D: Chapter/Section
            ws.cell(row=target_row, column=4).value = chapter_name
            # Column E: PDF/Word Page # (leave empty)
            # Column F: File Timecode
            ws.cell(row=target_row, column=6).value = timecode
            # Column G: Error Type
            ws.cell(row=target_row, column=7).value = error_type
            # Column H: Comments
            ws.cell(row=target_row, column=8).value = comments

            # Save workbook
            wb.save(crx_path)

            self.send_json_response({
                'success': True,
                'errorNumber': error_num,
                'crxFile': crx_filename,
                'timecode': timecode
            })

        except Exception as e:
            print(f"Error adding to CRX: {e}")
            print(traceback.format_exc())
            self.send_json_response({'error': str(e)}, 500)

    def send_json_response(self, data, status=200):
        """Helper to send JSON response"""
        self.send_response(status)
        self.send_header('Content-type', 'application/json')
        self.end_headers()
        self.wfile.write(json.dumps(data).encode())

    def log_message(self, format, *args):
        """Custom log format"""
        print(f"[{self.log_date_time_string()}] {format % args}")


def main():
    global BASE_DIR, PORT

    # Parse command-line arguments
    if len(sys.argv) > 1:
        BASE_DIR = Path(sys.argv[1])
    if len(sys.argv) > 2:
        PORT = int(sys.argv[2])

    if not BASE_DIR.exists():
        print(f"Error: Base directory does not exist: {BASE_DIR}")
        sys.exit(1)

    server_address = ('', PORT)
    httpd = HTTPServer(server_address, ValidationReportHandler)

    print(f"="*60)
    print(f"Validation Report Viewer")
    print(f"="*60)
    print(f"Serving reports from: {BASE_DIR}")
    print(f"Server running at: http://localhost:{PORT}")
    print(f"Press Ctrl+C to stop")
    print(f"="*60)

    try:
        httpd.serve_forever()
    except KeyboardInterrupt:
        print("\n\nShutting down server...")
        httpd.shutdown()


if __name__ == '__main__':
    main()
